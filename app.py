# app.py
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import streamlit as st
from pathlib import Path
import glob
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io

# ================================================
# تنظیمات اولیه و توابع کمکی
# ================================================

def setup_directories():
    """ایجاد پوشه‌های مورد نیاز اگر وجود نداشته باشند"""
    directories = ['history', 'reports', 'without_owner', 'personnel_reports', 'analysis_output', 'backups', 'data']
    for dir_name in directories:
        Path(dir_name).mkdir(exist_ok=True)
    return directories

def calculate_priority(urgency, importance):
    """محاسبه اولویت بر اساس ماتریس فوریت و اهمیت"""
    if pd.isna(urgency) or pd.isna(importance):
        return "نامشخص"
    
    try:
        urgency = int(urgency)
        importance = int(importance)
    except (ValueError, TypeError):
        return "نامشخص"
    
    if urgency == 1 and importance == 1:
        return "حیاتی"
    elif urgency == 0 and importance == 1:
        return "بسیار مهم"
    elif urgency == 1 and importance == 0:
        return "مهم"
    elif urgency == 0 and importance == 0:
        return "قابل بررسی"
    else:
        return "نامشخص"

def safe_convert_to_int(value, default=0):
    """تبدیل امن مقادیر به عدد صحیح"""
    if pd.isna(value) or value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default

def safe_convert_to_float(value, default=0.0):
    """تبدیل امن مقادیر به عدد اعشاری"""
    if pd.isna(value) or value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def convert_to_serializable(obj):
    """تبدیل انواع داده غیرقابل سریال‌سازی به انواع قابل سریال‌سازی"""
    if isinstance(obj, (np.int64, np.int32, np.int16, np.int8)):
        return int(obj)
    elif isinstance(obj, (np.float64, np.float32, np.float16)):
        return float(obj)
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, pd.Timestamp):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    elif pd.isna(obj):
        return None
    return obj

def get_next_id(df):
    """دریافت ID بعدی برای فعالیت جدید"""
    if 'ردیف' in df.columns:
        max_id = df['ردیف'].apply(lambda x: safe_convert_to_int(x, 0)).max()
        return max_id + 1
    return 1


def safe_parse_date(date_value, default_date=None):
    """تبدیل امن تاریخ با مدیریت مقادیر خالی و نامعتبر"""
    if default_date is None:
        default_date = datetime.now().date()
    
    if pd.isna(date_value) or date_value is None or date_value == '':
        return default_date
    
    try:
        # اگر از نوع datetime یا date است
        if isinstance(date_value, (datetime, pd.Timestamp)):
            return date_value.date()
        # اگر رشته است
        elif isinstance(date_value, str):
            return pd.to_datetime(date_value).date()
        else:
            return default_date
    except:
        return default_date

def safe_format_date(date_value, default_str=""):
    """فرمت‌بندی امن تاریخ برای نمایش"""
    if pd.isna(date_value) or date_value is None or date_value == '':
        return default_str
    
    try:
        if isinstance(date_value, (datetime, pd.Timestamp)):
            return date_value.strftime("%Y-%m-%d")
        elif isinstance(date_value, str):
            return date_value
        else:
            return default_str
    except:
        return default_str
# ================================================
# ساختار سلسله‌مراتب پوزیشن‌های سازمانی
# ================================================

def create_organizational_structure():
    """ایجاد ساختار سلسله‌مراتب پوزیشن‌های سازمانی"""
    structure = {
        "سطح رهبری": {
            "هیئت مدیره": [],
            "مدیرعامل": [],
            "مسئول دفتر عامل": [],
            "شورای نوآوری و فناوری": []
        },
        "سطح عملیاتی": {
            "مدیر فناوری و توسعه کسب و کار": {
                "مدیر فنی و ارزیابی": {
                    "رییس تیم شناسایی و ارزیابی فناوری": [
                        "کارشناس ارشد (شناسایی و ارزیابی فناوری)",
                        "کارشناس (شناسایی و ارزیابی فناوری)"
                    ],
                    "رییس تیم بازاریابی صنعتی": [
                        "کارشناس ارشد (بازاریابی صنعتی)",
                        "کارشناس (بازاریابی صنعتی)"
                    ]
                },
                "مدیر بازاریابی و فروش": {
                    "رییس تیم استاندارد و تجاری سازی تولید": [
                        "کارشناس ارشد (استاندارد و تجاری سازی تولید)",
                        "کارشناس (استاندارد و تجاری سازی تولید)"
                    ],
                    "رییس تیم فروش": [
                        "کارشناس ارشد (فروش)",
                        "کارشناس (فروش)"
                    ]
                }
            }
        },
        "سطح ستادی": {
            "مدیر پشتیبانی": {
                "رییس منابع انسانی و مالی": [
                    "کارشناس منابع انسانی",
                    "کارشناس مالی"
                ],
                "رییس سیستمها و روابط عمومی": [
                    "کارشناس ارشد (سیستمها و روابط عمومی)",
                    "کارشناس (سیستمها و روابط عمومی)",
                    "کارشناس ارشد حقوقی و پیمانها"
                ]
            }
        }
    }
    return structure

def extract_all_positions(structure=None, parent_path=""):
    """استخراج تمام پوزیشن‌ها از ساختار سلسله‌مراتبی"""
    if structure is None:
        structure = create_organizational_structure()
    
    positions = []
    
    def extract_recursive(node, level_path):
        if isinstance(node, dict):
            for key, value in node.items():
                current_path = f"{level_path} > {key}" if level_path else key
                positions.append({
                    'نام': key,
                    'مسیر کامل': current_path,
                    'سطح': len(current_path.split(' > ')) if level_path else 1,
                    'نوع': 'مدیریتی' if isinstance(value, dict) else 'عملیاتی'
                })
                extract_recursive(value, current_path)
        elif isinstance(node, list):
            for item in node:
                current_path = f"{level_path} > {item}" if level_path else item
                positions.append({
                    'نام': item,
                    'مسیر کامل': current_path,
                    'سطح': len(current_path.split(' > ')) if level_path else 1,
                    'نوع': 'کارشناسی'
                })
    
    extract_recursive(structure, "")
    return pd.DataFrame(positions)

def get_position_level(position_name, structure=None):
    """دریافت سطح سازمانی یک پوزیشن"""
    if structure is None:
        structure = create_organizational_structure()
    
    def search_recursive(node, target, current_level=1):
        if isinstance(node, dict):
            for key, value in node.items():
                if key == target:
                    return current_level
                result = search_recursive(value, target, current_level + 1)
                if result:
                    return result
        elif isinstance(node, list):
            if target in node:
                return current_level
        return None
    
    level = search_recursive(structure, position_name)
    return level if level else 0

def get_position_hierarchy(position_name, structure=None):
    """دریافت سلسله‌مراتب یک پوزیشن (از بالا به پایین)"""
    if structure is None:
        structure = create_organizational_structure()
    
    hierarchy = []
    
    def search_recursive(node, target, path=[]):
        if isinstance(node, dict):
            for key, value in node.items():
                current_path = path + [key]
                if key == target:
                    return current_path
                result = search_recursive(value, target, current_path)
                if result:
                    return result
        elif isinstance(node, list):
            if target in node:
                return path + [target]
        return None
    
    result = search_recursive(structure, position_name)
    return result if result else [position_name]

# ================================================
# توابع مدیریت چارت سازمانی (پرسنل و پوزیشن‌ها)
# ================================================

def load_organizational_chart():
    """بارگذاری چارت سازمانی از فایل JSON"""
    chart_file = 'data/organizational_chart.json'
    
    # ساختار پیش‌فرض با پوزیشن‌های سازمانی
    default_chart = {
        'personnel': [],  # لیست پرسنل
        'positions': extract_all_positions()['نام'].tolist(),  # لیست پوزیشن‌ها از ساختار
        'personnel_positions': {},  # نگاشت پرسنل به پوزیشن‌ها (چند به چند)
        'structure': create_organizational_structure()  # ساختار سلسله‌مراتبی
    }
    
    if os.path.exists(chart_file):
        try:
            with open(chart_file, 'r', encoding='utf-8') as f:
                chart_data = json.load(f)
                # اطمینان از وجود کلیدهای مورد نیاز
                if 'personnel' not in chart_data:
                    chart_data['personnel'] = []
                if 'positions' not in chart_data:
                    chart_data['positions'] = extract_all_positions()['نام'].tolist()
                if 'personnel_positions' not in chart_data:
                    chart_data['personnel_positions'] = {}
                if 'structure' not in chart_data:
                    chart_data['structure'] = create_organizational_structure()
                return chart_data
        except:
            return default_chart
    else:
        # اگر فایل وجود نداشت، از داده‌های موجود در DB استخراج کن
        try:
            df = pd.read_excel('DB.xlsx', header=0, dtype=str)
            personnel = set()
            
            if 'مسئول1' in df.columns:
                personnel.update(df['مسئول1'].dropna().unique())
            if 'مسئول2' in df.columns:
                personnel.update(df['مسئول2'].dropna().unique())
            
            # حذف مقادیر خالی
            personnel = {p for p in personnel if p and str(p).strip()}
            
            default_chart['personnel'] = sorted(list(personnel))
            
            # استخراج پوزیشن‌های استفاده شده در دیتابیس
            if 'پوزیشن_سازمانی' in df.columns:
                db_positions = set(df['پوزیشن_سازمانی'].dropna().unique())
                db_positions = {p for p in db_positions if p and str(p).strip()}
                
                # اضافه کردن پوزیشن‌های جدید به لیست اصلی
                all_positions = set(default_chart['positions'])
                all_positions.update(db_positions)
                default_chart['positions'] = sorted(list(all_positions))
            
            # ذخیره فایل
            save_organizational_chart(default_chart)
        except:
            pass
        
        return default_chart

def save_organizational_chart(chart_data):
    """ذخیره چارت سازمانی در فایل JSON"""
    chart_file = 'data/organizational_chart.json'
    
    # اطمینان از وجود پوشه data
    Path('data').mkdir(exist_ok=True)
    
    # مرتب‌سازی لیست‌ها
    if 'personnel' in chart_data:
        chart_data['personnel'] = sorted([p for p in chart_data['personnel'] if p and str(p).strip()])
    if 'positions' in chart_data:
        chart_data['positions'] = sorted([p for p in chart_data['positions'] if p and str(p).strip()])
    
    with open(chart_file, 'w', encoding='utf-8') as f:
        json.dump(chart_data, f, ensure_ascii=False, indent=4)

def add_personnel(chart_data, name):
    """افزودن پرسنل جدید به چارت"""
    if name and name.strip():
        name = name.strip()
        if name not in chart_data['personnel']:
            chart_data['personnel'].append(name)
            chart_data['personnel'] = sorted(chart_data['personnel'])
            
            # ایجاد ورود خالی برای پوزیشن‌های این پرسنل
            if name not in chart_data['personnel_positions']:
                chart_data['personnel_positions'][name] = []
            
            save_organizational_chart(chart_data)
            return True, f"پرسنل '{name}' با موفقیت اضافه شد."
        else:
            return False, f"پرسنل '{name}' قبلاً ثبت شده است."
    return False, "نام پرسنل نمی‌تواند خالی باشد."

def assign_position_to_personnel(chart_data, personnel_name, position_name):
    """اختصاص پوزیشن به پرسنل"""
    if personnel_name in chart_data['personnel'] and position_name in chart_data['positions']:
        if personnel_name not in chart_data['personnel_positions']:
            chart_data['personnel_positions'][personnel_name] = []
        
        if position_name not in chart_data['personnel_positions'][personnel_name]:
            chart_data['personnel_positions'][personnel_name].append(position_name)
            chart_data['personnel_positions'][personnel_name] = sorted(chart_data['personnel_positions'][personnel_name])
            save_organizational_chart(chart_data)
            return True, f"پوزیشن '{position_name}' به '{personnel_name}' اختصاص یافت."
        else:
            return False, f"این پوزیشن قبلاً به این پرسنل اختصاص یافته است."
    return False, "پرسنل یا پوزیشن نامعتبر است."

def remove_position_from_personnel(chart_data, personnel_name, position_name):
    """حذف پوزیشن از پرسنل"""
    if (personnel_name in chart_data['personnel_positions'] and 
        position_name in chart_data['personnel_positions'][personnel_name]):
        chart_data['personnel_positions'][personnel_name].remove(position_name)
        save_organizational_chart(chart_data)
        return True, f"پوزیشن '{position_name}' از '{personnel_name}' حذف شد."
    return False, "این پوزیشن برای این پرسنل یافت نشد."

def remove_personnel(chart_data, name):
    """حذف پرسنل از چارت"""
    if name in chart_data['personnel']:
        chart_data['personnel'].remove(name)
        if name in chart_data['personnel_positions']:
            del chart_data['personnel_positions'][name]
        save_organizational_chart(chart_data)
        return True, f"پرسنل '{name}' با موفقیت حذف شد."
    return False, f"پرسنل '{name}' یافت نشد."

def add_position(chart_data, position):
    """افزودن پوزیشن جدید به چارت"""
    if position and position.strip():
        position = position.strip()
        if position not in chart_data['positions']:
            chart_data['positions'].append(position)
            chart_data['positions'] = sorted(chart_data['positions'])
            save_organizational_chart(chart_data)
            return True, f"پوزیشن '{position}' با موفقیت اضافه شد."
        else:
            return False, f"پوزیشن '{position}' قبلاً ثبت شده است."
    return False, "عنوان پوزیشن نمی‌تواند خالی باشد."

def remove_position(chart_data, position):
    """حذف پوزیشن از چارت"""
    if position in chart_data['positions']:
        # حذف از تمام انتساب‌های پرسنل
        for personnel in chart_data['personnel_positions']:
            if position in chart_data['personnel_positions'][personnel]:
                chart_data['personnel_positions'][personnel].remove(position)
        
        chart_data['positions'].remove(position)
        save_organizational_chart(chart_data)
        return True, f"پوزیشن '{position}' با موفقیت حذف شد."
    return False, f"پوزیشن '{position}' یافت نشد."

def update_from_db(df, chart_data):
    """به‌روزرسانی چارت از دیتابیس فعالیت‌ها"""
    # استخراج پرسنل از دیتابیس
    db_personnel = set()
    if 'مسئول1' in df.columns:
        db_personnel.update(df['مسئول1'].dropna().unique())
    if 'مسئول2' in df.columns:
        db_personnel.update(df['مسئول2'].dropna().unique())
    
    # استخراج پوزیشن‌ها از دیتابیس
    db_positions = set()
    if 'پوزیشن_سازمانی' in df.columns:
        db_positions.update(df['پوزیشن_سازمانی'].dropna().unique())
    
    # حذف مقادیر خالی
    db_personnel = {p for p in db_personnel if p and str(p).strip()}
    db_positions = {p for p in db_positions if p and str(p).strip()}
    
    # اضافه کردن مقادیر جدید به چارت
    new_personnel = db_personnel - set(chart_data['personnel'])
    new_positions = db_positions - set(chart_data['positions'])
    
    changes = []
    
    if new_personnel:
        for person in new_personnel:
            chart_data['personnel'].append(person)
            if person not in chart_data['personnel_positions']:
                chart_data['personnel_positions'][person] = []
        changes.append(f"{len(new_personnel)} پرسنل جدید")
    
    if new_positions:
        chart_data['positions'].extend(list(new_positions))
        changes.append(f"{len(new_positions)} پوزیشن جدید")
    
    if changes:
        chart_data['personnel'] = sorted(chart_data['personnel'])
        chart_data['positions'] = sorted(chart_data['positions'])
        save_organizational_chart(chart_data)
        return True, new_personnel, new_positions
    
    return False, set(), set()

def get_personnel_activities_by_position(df, personnel_name, chart_data):
    """دریافت تعداد فعالیت‌های یک پرسنل در هر پوزیشن با جزئیات وضعیت"""
    mask = (df['مسئول1'] == personnel_name) | (df['مسئول2'] == personnel_name)
    person_activities = df[mask].copy()
    
    if person_activities.empty:
        return pd.DataFrame()
    
    result = []
    
    # پوزیشن‌های اختصاص داده شده به این پرسنل
    assigned_positions = chart_data['personnel_positions'].get(personnel_name, [])
    
    # اگر پوزیشنی اختصاص داده نشده، از پوزیشن‌های موجود در فعالیت‌ها استفاده کن
    if not assigned_positions:
        assigned_positions = person_activities['پوزیشن_سازمانی'].unique()
    
    for position in assigned_positions:
        if not position or not str(position).strip():
            continue
            
        # فعالیت‌های این پوزیشن
        pos_acts = person_activities[person_activities['پوزیشن_سازمانی'] == position]
        
        if pos_acts.empty:
            # اگر فعالیتی برای این پوزیشن وجود ندارد، رکورد با صفر ایجاد کن
            result.append({
                'پوزیشن': position,
                'سطح سازمانی': get_position_level(position, chart_data.get('structure')),
                'مسیر سلسله‌مراتب': ' > '.join(get_position_hierarchy(position, chart_data.get('structure'))),
                'تعداد کل': 0,
                'ToDo': 0,
                'Doing': 0,
                'Done': 0,
                'Blocked': 0,
                'درصد تکمیل': 0,
                'میانگین پیشرفت': 0
            })
        else:
            # آمار وضعیت‌ها
            todo_count = (pos_acts['وضعیت'] == 'ToDo').sum()
            doing_count = (pos_acts['وضعیت'] == 'Doing').sum()
            done_count = (pos_acts['وضعیت'] == 'Done').sum()
            blocked_count = (pos_acts['وضعیت'] == 'Blocked').sum()
            total = len(pos_acts)
            
            completion_rate = (done_count / total * 100) if total > 0 else 0
            
            result.append({
                'پوزیشن': position,
                'سطح سازمانی': get_position_level(position, chart_data.get('structure')),
                'مسیر سلسله‌مراتب': ' > '.join(get_position_hierarchy(position, chart_data.get('structure'))),
                'تعداد کل': total,
                'ToDo': todo_count,
                'Doing': doing_count,
                'Done': done_count,
                'Blocked': blocked_count,
                'درصد تکمیل': round(completion_rate, 1),
                'میانگین پیشرفت': round(pos_acts['درصد پیشرفت واقعی'].mean(), 1)
            })
    
    return pd.DataFrame(result)

# ================================================
# توابع بارگذاری و ذخیره داده‌ها
# ================================================

def load_data():
    """بارگذاری داده‌ها از فایل Excel"""
    try:
        df = pd.read_excel('DB.xlsx', header=0, dtype=str)
        
        # پاکسازی و آماده‌سازی داده‌ها
        df = df.dropna(how='all')
        
        # اطمینان از وجود ستون ردیف
        if 'ردیف' not in df.columns:
            df.insert(0, 'ردیف', range(1, len(df) + 1))
        
        # تبدیل ستون‌های عددی
        if 'درصد پیشرفت واقعی' in df.columns:
            df['درصد پیشرفت واقعی'] = df['درصد پیشرفت واقعی'].apply(safe_convert_to_int)
        
        if 'فوریت' in df.columns:
            df['فوریت'] = df['فوریت'].apply(lambda x: safe_convert_to_int(x, 0))
        
        if 'اهمیت' in df.columns:
            df['اهمیت'] = df['اهمیت'].apply(lambda x: safe_convert_to_int(x, 0))
        
        # محاسبه اولویت
        if 'فوریت' in df.columns and 'اهمیت' in df.columns:
            df['اولویت_محاسبه‌شده'] = df.apply(
                lambda row: calculate_priority(row.get('فوریت'), row.get('اهمیت')), 
                axis=1
            )
        
        # پر کردن مقادیر خالی
        text_columns = ['فعالیت', 'وضعیت', 'مسئول1', 'مسئول2', 'توضیحات', 'پوزیشن_سازمانی']
        for col in text_columns:
            if col in df.columns:
                df[col] = df[col].fillna('')
            else:
                df[col] = '' if col != 'وضعیت' else 'ToDo'
        
        # اضافه کردن ستون تاریخ ایجاد اگر وجود ندارد
        if 'تاریخ_ایجاد' not in df.columns:
            df['تاریخ_ایجاد'] = datetime.now().strftime("%Y-%m-%d")
        
        if 'تاریخ_آخرین_تغییر' not in df.columns:
            df['تاریخ_آخرین_تغییر'] = ''
        
        if 'تاریخ_Done' not in df.columns:
            df['تاریخ_Done'] = ''
        
        return df
    except FileNotFoundError:
        st.error("فایل DB.xlsx یافت نشد! لطفاً فایل را در مسیر اصلی قرار دهید.")
        return None
    except Exception as e:
        st.error(f"خطا در خواندن فایل: {e}")
        return None

def save_data(df):
    """ذخیره داده‌ها در فایل Excel با پشتیبان‌گیری خودکار"""
    # ایجاد پشتیبان قبل از ذخیره
    backup_file = f"backups/DB_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if os.path.exists('DB.xlsx'):
        try:
            pd.read_excel('DB.xlsx').to_excel(backup_file, index=False)
        except:
            pass
    
    # ذخیره فایل اصلی
    df.to_excel('DB.xlsx', index=False)
    
    # حذف پشتیبان‌های قدیمی (فقط 10 تای آخر نگه دار)
    backup_files = sorted(glob.glob('backups/DB_backup_*.xlsx'))
    if len(backup_files) > 10:
        for file in backup_files[:-10]:
            try:
                os.remove(file)
            except:
                pass

def get_unique_responsibles(df):
    """استخراج لیست منحصر‌به‌فرد مسئولین"""
    responsibles = set()
    
    if 'مسئول1' in df.columns:
        responsibles.update(df['مسئول1'].dropna().unique())
    if 'مسئول2' in df.columns:
        responsibles.update(df['مسئول2'].dropna().unique())
    
    responsibles = {r for r in responsibles if pd.notna(r) and str(r).strip() and r != ''}
    
    return sorted(list(responsibles))

def get_unique_positions(df):
    """استخراج لیست پوزیشن‌های سازمانی"""
    if 'پوزیشن_سازمانی' in df.columns:
        positions = df['پوزیشن_سازمانی'].dropna().unique()
        return sorted([p for p in positions if p and str(p).strip()])
    return []

def get_activities_for_responsible(df, responsible):
    """دریافت فعالیت‌های مربوط به یک مسئول خاص"""
    mask = (df['مسئول1'] == responsible) | (df['مسئول2'] == responsible)
    return df[mask].copy()

def get_activities_without_responsible(df):
    """دریافت فعالیت‌هایی که هیچ مسئولی ندارند"""
    mask = (pd.isna(df['مسئول1']) | (df['مسئول1'] == '')) & \
           (pd.isna(df['مسئول2']) | (df['مسئول2'] == ''))
    return df[mask].copy()

# ================================================
# توابع مدیریت تاریخچه پیشرفته
# ================================================

def save_to_history(action_type, activity_data, changes_dict, user="سیستم"):
    """ذخیره هر نوع تغییر در تاریخچه با جزئیات کامل - با رفع مشکل JSON serialization"""
    history_file = 'history/History.xlsx'
    
    # تبدیل مقادیر به فرمت قابل سریال‌سازی
    serializable_changes = {}
    for key, value in changes_dict.items():
        serializable_changes[key] = convert_to_serializable(value)
    
    # ایجاد رکورد جدید
    new_record = pd.DataFrame([{
        'تاریخ': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'نوع_عملیات': action_type,
        'ردیف_فعالیت': convert_to_serializable(activity_data.get('ردیف', '')),
        'فعالیت': activity_data.get('فعالیت', ''),
        'مسئول1': activity_data.get('مسئول1', ''),
        'مسئول2': activity_data.get('مسئول2', ''),
        'پوزیشن': activity_data.get('پوزیشن_سازمانی', ''),
        'اولویت': activity_data.get('اولویت_محاسبه‌شده', 'نامشخص'),
        'وضعیت_قدیم': convert_to_serializable(changes_dict.get('وضعیت_قدیم', '')),
        'وضعیت_جدید': convert_to_serializable(changes_dict.get('وضعیت_جدید', '')),
        'مسئول_قدیم': convert_to_serializable(changes_dict.get('مسئول_قدیم', '')),
        'مسئول_جدید': convert_to_serializable(changes_dict.get('مسئول_جدید', '')),
        'پیشرفت_قدیم': convert_to_serializable(changes_dict.get('پیشرفت_قدیم', '')),
        'پیشرفت_جدید': convert_to_serializable(changes_dict.get('پیشرفت_جدید', '')),
        'توضیحات_قدیم': convert_to_serializable(changes_dict.get('توضیحات_قدیم', '')),
        'توضیحات_جدید': convert_to_serializable(changes_dict.get('توضیحات_جدید', '')),
        'اولویت_قدیم': convert_to_serializable(changes_dict.get('اولویت_قدیم', '')),
        'اولویت_جدید': convert_to_serializable(changes_dict.get('اولویت_جدید', '')),
        'کاربر': user,
        'جزئیات_کامل': json.dumps(serializable_changes, ensure_ascii=False)
    }])
    
    if os.path.exists(history_file):
        existing_history = pd.read_excel(history_file)
        updated_history = pd.concat([existing_history, new_record], ignore_index=True)
    else:
        updated_history = new_record
    
    updated_history.to_excel(history_file, index=False)
    return updated_history

def get_activity_history(activity_id):
    """دریافت تاریخچه یک فعالیت خاص"""
    history_file = 'history/History.xlsx'
    if not os.path.exists(history_file):
        return pd.DataFrame()
    
    history_df = pd.read_excel(history_file)
    if 'ردیف_فعالیت' in history_df.columns:
        activity_id_str = str(convert_to_serializable(activity_id))
        return history_df[history_df['ردیف_فعالیت'].astype(str) == activity_id_str].sort_values('تاریخ', ascending=False)
    return pd.DataFrame()

def calculate_lead_time(history_df, activity_row):
    """محاسبه مدت زمان انجام فعالیت (از ToDo تا Done)"""
    if history_df.empty:
        return None
    
    # پیدا کردن اولین ثبت با وضعیت ToDo
    todo_records = history_df[history_df['وضعیت_جدید'] == 'ToDo']
    if not todo_records.empty:
        start_date = pd.to_datetime(todo_records.iloc[-1]['تاریخ'])
    else:
        # اگر در تاریخچه نبود، از تاریخ ایجاد استفاده کن
        start_date = pd.to_datetime(activity_row.get('تاریخ_ایجاد', datetime.now()))
    
    # پیدا کردن اولین ثبت با وضعیت Done
    done_records = history_df[history_df['وضعیت_جدید'] == 'Done']
    if not done_records.empty:
        end_date = pd.to_datetime(done_records.iloc[0]['تاریخ'])
        lead_time = end_date - start_date
        return lead_time.days
    return None

# ================================================
# توابع ایجاد و ویرایش فعالیت (با استفاده از چارت)
# ================================================

def add_new_activity(df, chart_data):
    """افزودن فعالیت جدید به دیتابیس با استفاده از لیست پرسنل و پوزیشن‌ها"""
    st.subheader("➕ ثبت فعالیت جدید")
    
    with st.form("new_activity_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            new_activity = st.text_input("عنوان فعالیت *", key="new_act")
            
            # انتخاب مسئول اول از لیست پرسنل
            personnel_list = [''] + chart_data['personnel']
            new_responsible1 = st.selectbox(
                "مسئول اول", 
                options=personnel_list,
                key="new_res1"
            )
            
            # انتخاب مسئول دوم از لیست پرسنل
            new_responsible2 = st.selectbox(
                "مسئول دوم", 
                options=personnel_list,
                key="new_res2"
            )
            
            # انتخاب پوزیشن سازمانی از لیست
            position_list = [''] + chart_data['positions']
            new_position = st.selectbox(
                "پوزیشن سازمانی", 
                options=position_list,
                key="new_pos"
            )
        
        with col2:
            new_status = st.selectbox("وضعیت *", ["ToDo", "Doing", "Done", "Blocked"], key="new_status")
            new_urgency = st.selectbox("فوریت (0=غیرفوری، 1=فوری)", [0, 1], key="new_urg")
            new_importance = st.selectbox("اهمیت (0=معمولی، 1=مهم)", [0, 1], key="new_imp")
            new_progress = st.slider("درصد پیشرفت", 0, 100, 0, key="new_prog")
            new_start_date = st.date_input("تاریخ شروع", datetime.now(), key="new_start")
            new_end_date = st.date_input("تاریخ پایان", datetime.now() + timedelta(days=7), key="new_end")
        
        new_description = st.text_area("توضیحات", key="new_desc")
        
        submitted = st.form_submit_button("✅ ثبت فعالیت جدید", type="primary")
        
        if submitted:
            if not new_activity:
                st.error("عنوان فعالیت الزامی است!")
                return df
            
            # ایجاد ردیف جدید
            new_row = {
                'ردیف': get_next_id(df),
                'فعالیت': new_activity,
                'وضعیت': new_status,
                'مسئول1': new_responsible1,
                'مسئول2': new_responsible2,
                'پوزیشن_سازمانی': new_position,
                'فوریت': new_urgency,
                'اهمیت': new_importance,
                'درصد پیشرفت واقعی': new_progress,
                'تاریخ شروع': new_start_date.strftime("%Y-%m-%d"),
                'تاریخ پایان': new_end_date.strftime("%Y-%m-%d"),
                'توضیحات': new_description,
                'تاریخ_ایجاد': datetime.now().strftime("%Y-%m-%d"),
                'تاریخ_آخرین_تغییر': datetime.now().strftime("%Y-%m-%d"),
                'تاریخ_Done': datetime.now().strftime("%Y-%m-%d") if new_status == 'Done' else ''
            }
            
            # محاسبه اولویت
            new_row['اولویت_محاسبه‌شده'] = calculate_priority(new_urgency, new_importance)
            
            # اضافه کردن به دیتافریم
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
            # ذخیره در تاریخچه
            save_to_history(
                action_type="ایجاد",
                activity_data=new_row,
                changes_dict={"ایجاد_فعالیت": new_activity, "وضعیت_اولیه": new_status}
            )
            
            # ذخیره در فایل
            save_data(df)
            
            # به‌روزرسانی چارت از دیتابیس
            update_from_db(df, chart_data)
            
            st.success(f"✅ فعالیت '{new_activity}' با موفقیت ایجاد شد!")
            st.balloons()
            
            return df
    
    return df

def edit_activity(df, chart_data):
    """ویرایش مقادیر یک فعالیت موجود با قابلیت فیلتر"""
    st.subheader("✏️ ویرایش فعالیت")
    
    if len(df) == 0:
        st.warning("هیچ فعالیتی برای ویرایش وجود ندارد!")
        return df
    
    # ================================================
    # فیلترهای مخصوص صفحه ویرایش
    # ================================================
    st.markdown("### 🔍 فیلترهای جستجو و تفکیک فعالیت‌ها")
    
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    
    with col_f1:
        # فیلتر بر اساس مسئول
        all_responsibles = ['همه'] + chart_data['personnel']
        selected_resp_filter = st.selectbox(
            "فیلتر بر اساس مسئول",
            options=all_responsibles,
            key="edit_resp_filter"
        )
    
    with col_f2:
        # فیلتر بر اساس وضعیت
        all_statuses = ['همه'] + [s for s in df['وضعیت'].dropna().unique() if s]
        selected_status_filter = st.selectbox(
            "فیلتر بر اساس وضعیت",
            options=all_statuses,
            key="edit_status_filter"
        )
    
    with col_f3:
        # فیلتر بر اساس اولویت
        all_priorities = ['همه', 'حیاتی', 'بسیار مهم', 'مهم', 'قابل بررسی', 'نامشخص']
        selected_priority_filter = st.selectbox(
            "فیلتر بر اساس اولویت",
            options=all_priorities,
            key="edit_priority_filter"
        )
    
    with col_f4:
        # فیلتر متن برای جستجو در عنوان فعالیت
        search_text = st.text_input("جستجو در عنوان فعالیت", key="edit_search")
    
    # ================================================
    # اعمال فیلترها روی لیست فعالیت‌ها
    # ================================================
    filtered_df = df.copy()
    
    if selected_resp_filter != 'همه':
        filtered_df = filtered_df[
            (filtered_df['مسئول1'] == selected_resp_filter) | 
            (filtered_df['مسئول2'] == selected_resp_filter)
        ]
    
    if selected_status_filter != 'همه':
        filtered_df = filtered_df[filtered_df['وضعیت'] == selected_status_filter]
    
    if selected_priority_filter != 'همه':
        filtered_df = filtered_df[filtered_df['اولویت_محاسبه‌شده'] == selected_priority_filter]
    
    if search_text:
        filtered_df = filtered_df[filtered_df['فعالیت'].str.contains(search_text, case=False, na=False)]
    
    # ================================================
    # نمایش آمار و لیست فیلتر شده
    # ================================================
    st.markdown(f"**تعداد فعالیت‌های یافت شده:** {len(filtered_df)}")
    
    if len(filtered_df) > 0:
        # نمایش لیست فیلتر شده به صورت خلاصه
        with st.expander("📋 نمایش لیست فعالیت‌های فیلتر شده", expanded=False):
            display_cols = ['ردیف', 'فعالیت', 'وضعیت', 'مسئول1', 'مسئول2', 'پوزیشن_سازمانی', 
                           'اولویت_محاسبه‌شده', 'درصد پیشرفت واقعی', 'تاریخ شروع', 'تاریخ پایان']
            available_display = [col for col in display_cols if col in filtered_df.columns]
            st.dataframe(filtered_df[available_display], use_container_width=True)
        
        st.markdown("---")
        
        # انتخاب فعالیت برای ویرایش از لیست فیلتر شده
        activities_list = filtered_df['فعالیت'].tolist()
        
        if activities_list:
            selected_activity = st.selectbox(
                "فعالیت مورد نظر برای ویرایش را انتخاب کنید", 
                activities_list, 
                key="edit_select"
            )
            
            if selected_activity:
                # پیدا کردن ایندکس فعالیت در دیتافریم اصلی (نه فیلتر شده)
                activity_idx = df[df['فعالیت'] == selected_activity].index[0]
                activity_row = df.loc[activity_idx]
                
                st.info(f"**در حال ویرایش:** {selected_activity}")
                
                # نمایش مقادیر فعلی
                with st.expander("📋 مقادیر فعلی", expanded=True):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write("**وضعیت:**", activity_row.get('وضعیت', ''))
                        st.write("**مسئول اول:**", activity_row.get('مسئول1', ''))
                        st.write("**مسئول دوم:**", activity_row.get('مسئول2', ''))
                        st.write("**پوزیشن:**", activity_row.get('پوزیشن_سازمانی', ''))
                    with col2:
                        st.write("**فوریت:**", activity_row.get('فوریت', ''))
                        st.write("**اهمیت:**", activity_row.get('اهمیت', ''))
                        st.write("**اولویت:**", activity_row.get('اولویت_محاسبه‌شده', ''))
                        st.write("**پیشرفت:**", activity_row.get('درصد پیشرفت واقعی', ''), "%")
                        
                        # نمایش امن تاریخ‌ها
                        start_date_val = safe_format_date(activity_row.get('تاریخ شروع'), 'تنظیم نشده')
                        end_date_val = safe_format_date(activity_row.get('تاریخ پایان'), 'تنظیم نشده')
                        st.write("**تاریخ شروع:**", start_date_val)
                        st.write("**تاریخ پایان:**", end_date_val)
                    st.write("**توضیحات:**", activity_row.get('توضیحات', ''))
                
                # فرم ویرایش
                with st.form("edit_activity_form"):
                    st.markdown("---")
                    st.subheader("مقادیر جدید")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        new_status = st.selectbox(
                            "وضعیت جدید",
                            ["ToDo", "Doing", "Done", "Blocked"],
                            index=["ToDo", "Doing", "Done", "Blocked"].index(activity_row.get('وضعیت', 'ToDo')) 
                            if activity_row.get('وضعیت', 'ToDo') in ["ToDo", "Doing", "Done", "Blocked"] else 0
                        )
                        
                        # انتخاب مسئول اول از لیست پرسنل
                        personnel_list = [''] + chart_data['personnel']
                        default_res1 = activity_row.get('مسئول1', '')
                        default_res1_index = personnel_list.index(default_res1) if default_res1 in personnel_list else 0
                        new_responsible1 = st.selectbox(
                            "مسئول اول جدید",
                            options=personnel_list,
                            index=default_res1_index,
                            key="edit_res1"
                        )
                        
                        # انتخاب مسئول دوم از لیست پرسنل
                        default_res2 = activity_row.get('مسئول2', '')
                        default_res2_index = personnel_list.index(default_res2) if default_res2 in personnel_list else 0
                        new_responsible2 = st.selectbox(
                            "مسئول دوم جدید",
                            options=personnel_list,
                            index=default_res2_index,
                            key="edit_res2"
                        )
                        
                        # انتخاب پوزیشن از لیست
                        position_list = [''] + chart_data['positions']
                        default_pos = activity_row.get('پوزیشن_سازمانی', '')
                        default_pos_index = position_list.index(default_pos) if default_pos in position_list else 0
                        new_position = st.selectbox(
                            "پوزیشن سازمانی جدید",
                            options=position_list,
                            index=default_pos_index,
                            key="edit_pos"
                        )
                        
                        # تاریخ شروع با مدیریت خطا
                        current_start = safe_parse_date(activity_row.get('تاریخ شروع'))
                        new_start_date = st.date_input(
                            "تاریخ شروع جدید",
                            value=current_start,
                            key="edit_start_date"
                        )
                    
                    with col2:
                        new_urgency = st.selectbox(
                            "فوریت جدید",
                            [0, 1],
                            index=0 if activity_row.get('فوریت', 0) == 0 else 1
                        )
                        
                        new_importance = st.selectbox(
                            "اهمیت جدید",
                            [0, 1],
                            index=0 if activity_row.get('اهمیت', 0) == 0 else 1
                        )
                        
                        new_progress = st.slider(
                            "درصد پیشرفت جدید",
                            0, 100,
                            value=int(activity_row.get('درصد پیشرفت واقعی', 0))
                        )
                        
                        # تاریخ پایان با مدیریت خطا
                        current_end = safe_parse_date(activity_row.get('تاریخ پایان'), datetime.now().date() + timedelta(days=7))
                        new_end_date = st.date_input(
                            "تاریخ پایان جدید",
                            value=current_end,
                            key="edit_end_date"
                        )
                    
                    new_description = st.text_area("توضیحات جدید", value=activity_row.get('توضیحات', ''))
                    
                    # دکمه ثبت تغییرات
                    submitted = st.form_submit_button("💾 ثبت تغییرات", type="primary")
                    
                    if submitted:
                        # جمع‌آوری تغییرات برای تاریخچه
                        changes = {}
                        
                        # بررسی تغییرات
                        if new_status != activity_row.get('وضعیت', ''):
                            changes['وضعیت_قدیم'] = activity_row.get('وضعیت', '')
                            changes['وضعیت_جدید'] = new_status
                            df.loc[activity_idx, 'وضعیت'] = new_status
                            
                            # اگر وضعیت Done شد، تاریخ Done را ثبت کن
                            if new_status == 'Done':
                                df.loc[activity_idx, 'تاریخ_Done'] = datetime.now().strftime("%Y-%m-%d")
                        
                        if new_responsible1 != activity_row.get('مسئول1', ''):
                            changes['مسئول_قدیم'] = f"مسئول1: {activity_row.get('مسئول1', '')}"
                            changes['مسئول_جدید'] = f"مسئول1: {new_responsible1}"
                            df.loc[activity_idx, 'مسئول1'] = new_responsible1
                        
                        if new_responsible2 != activity_row.get('مسئول2', ''):
                            changes['مسئول2_قدیم'] = activity_row.get('مسئول2', '')
                            changes['مسئول2_جدید'] = new_responsible2
                            df.loc[activity_idx, 'مسئول2'] = new_responsible2
                        
                        if new_position != activity_row.get('پوزیشن_سازمانی', ''):
                            changes['پوزیشن_قدیم'] = activity_row.get('پوزیشن_سازمانی', '')
                            changes['پوزیشن_جدید'] = new_position
                            df.loc[activity_idx, 'پوزیشن_سازمانی'] = new_position
                        
                        if new_urgency != activity_row.get('فوریت', 0):
                            changes['فوریت_قدیم'] = activity_row.get('فوریت', 0)
                            changes['فوریت_جدید'] = new_urgency
                            df.loc[activity_idx, 'فوریت'] = new_urgency
                        
                        if new_importance != activity_row.get('اهمیت', 0):
                            changes['اهمیت_قدیم'] = activity_row.get('اهمیت', 0)
                            changes['اهمیت_جدید'] = new_importance
                            df.loc[activity_idx, 'اهمیت'] = new_importance
                        
                        if new_progress != activity_row.get('درصد پیشرفت واقعی', 0):
                            changes['پیشرفت_قدیم'] = activity_row.get('درصد پیشرفت واقعی', 0)
                            changes['پیشرفت_جدید'] = new_progress
                            df.loc[activity_idx, 'درصد پیشرفت واقعی'] = new_progress
                        
                        # بررسی تغییرات تاریخ
                        old_start = safe_format_date(activity_row.get('تاریخ شروع'))
                        new_start_str = new_start_date.strftime("%Y-%m-%d")
                        if old_start != new_start_str:
                            changes['تاریخ_شروع_قدیم'] = old_start
                            changes['تاریخ_شروع_جدید'] = new_start_str
                            df.loc[activity_idx, 'تاریخ شروع'] = new_start_str
                        
                        old_end = safe_format_date(activity_row.get('تاریخ پایان'))
                        new_end_str = new_end_date.strftime("%Y-%m-%d")
                        if old_end != new_end_str:
                            changes['تاریخ_پایان_قدیم'] = old_end
                            changes['تاریخ_پایان_جدید'] = new_end_str
                            df.loc[activity_idx, 'تاریخ پایان'] = new_end_str
                        
                        if new_description != activity_row.get('توضیحات', ''):
                            changes['توضیحات_قدیم'] = activity_row.get('توضیحات', '')
                            changes['توضیحات_جدید'] = new_description
                            df.loc[activity_idx, 'توضیحات'] = new_description
                        
                        # محاسبه مجدد اولویت اگر فوریت یا اهمیت تغییر کرده باشد
                        if 'فوریت_جدید' in changes or 'اهمیت_جدید' in changes:
                            new_priority = calculate_priority(new_urgency, new_importance)
                            old_priority = activity_row.get('اولویت_محاسبه‌شده', '')
                            if new_priority != old_priority:
                                changes['اولویت_قدیم'] = old_priority
                                changes['اولویت_جدید'] = new_priority
                                df.loc[activity_idx, 'اولویت_محاسبه‌شده'] = new_priority
                        
                        # به‌روزرسانی تاریخ آخرین تغییر
                        df.loc[activity_idx, 'تاریخ_آخرین_تغییر'] = datetime.now().strftime("%Y-%m-%d")
                        
                        if changes:
                            # ذخیره در تاریخچه
                            save_to_history(
                                action_type="ویرایش",
                                activity_data=activity_row.to_dict(),
                                changes_dict=changes
                            )
                            
                            # ذخیره در فایل
                            save_data(df)
                            
                            # به‌روزرسانی چارت از دیتابیس
                            update_from_db(df, chart_data)
                            
                            st.success("✅ تغییرات با موفقیت ثبت شد!")
                            
                            # نمایش خلاصه تغییرات
                            st.info("**تغییرات اعمال شده:**")
                            for key, value in changes.items():
                                if 'قدیم' in key and 'جدید' in key.replace('قدیم', 'جدید') in changes:
                                    continue
                                st.write(f"- {key}: {value}")
                        else:
                            st.info("هیچ تغییری اعمال نشد.")
                
                # نمایش تاریخچه فعالیت
                with st.expander("📚 تاریخچه تغییرات این فعالیت"):
                    activity_history = get_activity_history(activity_row.get('ردیف', ''))
                    if not activity_history.empty:
                        st.dataframe(activity_history[['تاریخ', 'نوع_عملیات', 'وضعیت_قدیم', 'وضعیت_جدید', 
                                                       'پیشرفت_قدیم', 'پیشرفت_جدید', 'کاربر']])
                    else:
                        st.write("تاریخچه‌ای برای این فعالیت یافت نشد.")
        else:
            st.warning("هیچ فعالیتی با فیلترهای انتخاب شده یافت نشد!")
    else:
        st.warning("هیچ فعالیتی با فیلترهای انتخاب شده یافت نشد!")
    
    return df
# ================================================
# صفحه مدیریت چارت سازمانی (پیشرفته)
# ================================================

def organizational_chart_page(df, chart_data):
    """صفحه مدیریت چارت سازمانی (پرسنل و پوزیشن‌ها) با ساختار سلسله‌مراتبی"""
    st.header("🏢 چارت سازمانی - ساختار سلسله‌مراتبی")
    
    # به‌روزرسانی خودکار از دیتابیس
    updated, new_personnel, new_positions = update_from_db(df, chart_data)
    if updated:
        if new_personnel:
            st.info(f"✅ {len(new_personnel)} پرسنل جدید از دیتابیس به چارت اضافه شد: {', '.join(new_personnel)}")
        if new_positions:
            st.info(f"✅ {len(new_positions)} پوزیشن جدید از دیتابیس به چارت اضافه شد: {', '.join(new_positions)}")
    
    # ایجاد تب‌های مختلف برای مدیریت
    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 ساختار سازمانی (درختی)", 
        "👥 مدیریت پرسنل", 
        "📋 مدیریت پوزیشن‌ها",
        "📊 تحلیل پیشرفته پرسنل-پوزیشن"
    ])
    
    with tab1:
        st.subheader("ساختار سلسله‌مراتبی سازمان")
        
        # نمایش ساختار درختی
        def display_tree(node, level=0):
            if isinstance(node, dict):
                for key, value in node.items():
                    st.markdown("&nbsp;&nbsp;" * level + f"📁 **{key}**")
                    display_tree(value, level + 1)
            elif isinstance(node, list):
                for item in node:
                    st.markdown("&nbsp;&nbsp;" * level + f"📄 {item}")
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("### سطوح مدیریتی")
            structure = chart_data.get('structure', create_organizational_structure())
            display_tree(structure)
        
        with col2:
            st.markdown("### لیست تمام پوزیشن‌ها")
            positions_df = extract_all_positions(structure)
            if not positions_df.empty:
                st.dataframe(positions_df, use_container_width=True)
    
    with tab2:
        st.subheader("مدیریت پرسنل")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            # نمایش لیست پرسنل با پوزیشن‌هایشان
            if chart_data['personnel']:
                personnel_data = []
                for person in chart_data['personnel']:
                    positions = chart_data['personnel_positions'].get(person, [])
                    positions_str = '، '.join(positions) if positions else 'بدون پوزیشن'
                    
                    # تعداد فعالیت‌های این پرسنل
                    person_acts = get_activities_for_responsible(df, person)
                    act_count = len(person_acts)
                    
                    personnel_data.append({
                        'نام پرسنل': person,
                        'پوزیشن‌ها': positions_str,
                        'تعداد فعالیت': act_count,
                        'تعداد پوزیشن': len(positions)
                    })
                
                personnel_df = pd.DataFrame(personnel_data)
                st.dataframe(personnel_df, use_container_width=True)
            else:
                st.warning("هیچ پرسنلی ثبت نشده است.")
        
        with col2:
            # فرم افزودن پرسنل جدید
            st.subheader("➕ افزودن پرسنل جدید")
            new_personnel = st.text_input("نام پرسنل", key="new_personnel_input")
            
            if st.button("افزودن به لیست", key="add_personnel_btn_unique"):
                success, message = add_personnel(chart_data, new_personnel)
                if success:
                    st.success(message)
                    st.rerun()
                else:
                    st.error(message)
            
            st.markdown("---")
            
            # فرم اختصاص پوزیشن به پرسنل
            st.subheader("🔗 اختصاص پوزیشن به پرسنل")
            
            if chart_data['personnel'] and chart_data['positions']:
                assign_person = st.selectbox(
                    "انتخاب پرسنل",
                    options=chart_data['personnel'],
                    key="assign_person_select"
                )
                
                assign_position = st.selectbox(
                    "انتخاب پوزیشن",
                    options=chart_data['positions'],
                    key="assign_position_select"
                )
                
                if st.button("اختصاص پوزیشن", key="assign_position_btn_unique"):
                    success, message = assign_position_to_personnel(chart_data, assign_person, assign_position)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
                
                # فرم حذف پوزیشن از پرسنل
                st.markdown("---")
                st.subheader("➖ حذف پوزیشن از پرسنل")
                
                if assign_person in chart_data['personnel_positions']:
                    person_positions = chart_data['personnel_positions'][assign_person]
                    if person_positions:
                        remove_position_from = st.selectbox(
                            "انتخاب پوزیشن برای حذف",
                            options=person_positions,
                            key="remove_position_from_person_select"
                        )
                        
                        if st.button("حذف پوزیشن", key="remove_position_from_person_btn"):
                            success, message = remove_position_from_personnel(chart_data, assign_person, remove_position_from)
                            if success:
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)
            
            st.markdown("---")
            
            # فرم حذف پرسنل
            st.subheader("➖ حذف پرسنل")
            if chart_data['personnel']:
                person_to_remove = st.selectbox(
                    "انتخاب پرسنل برای حذف",
                    options=chart_data['personnel'],
                    key="remove_personnel_select"
                )
                
                if st.button("حذف از لیست", key="remove_personnel_btn_unique"):
                    # بررسی اینکه آیا این پرسنل در فعالیت‌ها استفاده شده است
                    person_activities = get_activities_for_responsible(df, person_to_remove)
                    if not person_activities.empty:
                        st.warning(f"⚠️ این پرسنل در {len(person_activities)} فعالیت به عنوان مسئول ثبت شده است. در صورت حذف، این فعالیت‌ها بدون مسئول خواهند ماند.")
                    
                    success, message = remove_personnel(chart_data, person_to_remove)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
    
    with tab3:
        st.subheader("مدیریت پوزیشن‌های سازمانی")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            # نمایش لیست پوزیشن‌ها
            if chart_data['positions']:
                positions_df = pd.DataFrame({
                    'ردیف': range(1, len(chart_data['positions']) + 1),
                    'عنوان پوزیشن': chart_data['positions'],
                    'سطح سازمانی': [get_position_level(p, chart_data.get('structure')) for p in chart_data['positions']]
                })
                st.dataframe(positions_df, use_container_width=True)
            else:
                st.warning("هیچ پوزیشنی ثبت نشده است.")
        
        with col2:
            # فرم افزودن پوزیشن جدید
            st.subheader("➕ افزودن پوزیشن جدید")
            new_position = st.text_input("عنوان پوزیشن", key="new_position_input")
            
            if st.button("افزودن به لیست", key="add_position_btn_unique"):
                success, message = add_position(chart_data, new_position)
                if success:
                    st.success(message)
                    st.rerun()
                else:
                    st.error(message)
            
            st.markdown("---")
            
            # فرم حذف پوزیشن
            st.subheader("➖ حذف پوزیشن")
            if chart_data['positions']:
                position_to_remove = st.selectbox(
                    "انتخاب پوزیشن برای حذف",
                    options=chart_data['positions'],
                    key="remove_position_select"
                )
                
                if st.button("حذف از لیست", key="remove_position_btn_unique"):
                    # بررسی اینکه آیا این پوزیشن در فعالیت‌ها استفاده شده است
                    position_activities = df[df['پوزیشن_سازمانی'] == position_to_remove]
                    if not position_activities.empty:
                        st.warning(f"⚠️ این پوزیشن در {len(position_activities)} فعالیت استفاده شده است.")
                    
                    success, message = remove_position(chart_data, position_to_remove)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
    
    with tab4:
        st.subheader("تحلیل پیشرفته توزیع پرسنل در پوزیشن‌ها")
        
        if chart_data['personnel']:
            selected_person = st.selectbox(
                "انتخاب پرسنل برای تحلیل",
                options=chart_data['personnel'],
                key="analyze_person_select"
            )
            
            if selected_person:
                person_stats = get_personnel_activities_by_position(df, selected_person, chart_data)
                
                if not person_stats.empty:
                    st.markdown(f"### آمار فعالیت‌های {selected_person} به تفکیک پوزیشن")
                    
                    # نمایش جدول
                    st.dataframe(person_stats, use_container_width=True)
                    
                    # نمودارهای پیشرفته
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # نمودار توزیع فعالیت‌ها بر اساس وضعیت
                        status_data = []
                        for _, row in person_stats.iterrows():
                            if row['تعداد کل'] > 0:
                                status_data.append({
                                    'پوزیشن': row['پوزیشن'],
                                    'وضعیت': 'ToDo',
                                    'تعداد': row['ToDo']
                                })
                                status_data.append({
                                    'پوزیشن': row['پوزیشن'],
                                    'وضعیت': 'Doing',
                                    'تعداد': row['Doing']
                                })
                                status_data.append({
                                    'پوزیشن': row['پوزیشن'],
                                    'وضعیت': 'Done',
                                    'تعداد': row['Done']
                                })
                                status_data.append({
                                    'پوزیشن': row['پوزیشن'],
                                    'وضعیت': 'Blocked',
                                    'تعداد': row['Blocked']
                                })
                        
                        if status_data:
                            status_df = pd.DataFrame(status_data)
                            fig1 = px.bar(
                                status_df,
                                x='پوزیشن',
                                y='تعداد',
                                color='وضعیت',
                                title=f'توزیع وضعیت فعالیت‌های {selected_person} به تفکیک پوزیشن',
                                barmode='stack',
                                color_discrete_map={
                                    'ToDo': '#FFA07A',
                                    'Doing': '#87CEEB',
                                    'Done': '#98FB98',
                                    'Blocked': '#FF6347'
                                }
                            )
                            st.plotly_chart(fig1, use_container_width=True)
                    
                    with col2:
                        # نمودار درصد تکمیل و میانگین پیشرفت
                        fig2 = go.Figure()
                        fig2.add_trace(go.Bar(
                            x=person_stats['پوزیشن'],
                            y=person_stats['درصد تکمیل'],
                            name='درصد تکمیل شده',
                            marker_color='lightgreen',
                            text=person_stats['درصد تکمیل'].apply(lambda x: f'{x}%')
                        ))
                        fig2.add_trace(go.Bar(
                            x=person_stats['پوزیشن'],
                            y=person_stats['میانگین پیشرفت'],
                            name='میانگین پیشرفت',
                            marker_color='lightblue',
                            text=person_stats['میانگین پیشرفت'].apply(lambda x: f'{x}%')
                        ))
                        fig2.update_layout(
                            title='درصد تکمیل و میانگین پیشرفت به تفکیک پوزیشن',
                            barmode='group',
                            yaxis_title='درصد'
                        )
                        st.plotly_chart(fig2, use_container_width=True)
                    
                    # نمودار راداری برای مقایسه پوزیشن‌ها
                    if len(person_stats) > 2:
                        fig3 = go.Figure()
                        
                        categories = person_stats['پوزیشن'].tolist()
                        
                        fig3.add_trace(go.Scatterpolar(
                            r=person_stats['درصد تکمیل'].tolist(),
                            theta=categories,
                            fill='toself',
                            name='درصد تکمیل'
                        ))
                        
                        fig3.add_trace(go.Scatterpolar(
                            r=person_stats['میانگین پیشرفت'].tolist(),
                            theta=categories,
                            fill='toself',
                            name='میانگین پیشرفت'
                        ))
                        
                        fig3.update_layout(
                            polar=dict(
                                radialaxis=dict(
                                    visible=True,
                                    range=[0, 100]
                                )),
                            showlegend=True,
                            title='نمودار راداری مقایسه پوزیشن‌ها'
                        )
                        
                        st.plotly_chart(fig3, use_container_width=True)
                    
                    # تحلیل سلسله‌مراتبی
                    st.subheader("تحلیل سلسله‌مراتبی")
                    
                    # گروه‌بندی بر اساس سطح سازمانی
                    level_stats = person_stats.groupby('سطح سازمانی').agg({
                        'تعداد کل': 'sum',
                        'ToDo': 'sum',
                        'Doing': 'sum',
                        'Done': 'sum',
                        'Blocked': 'sum'
                    }).reset_index()
                    
                    level_stats['درصد تکمیل'] = (level_stats['Done'] / level_stats['تعداد کل'] * 100).round(1)
                    
                    fig4 = px.bar(level_stats, x='سطح سازمانی', y=['ToDo', 'Doing', 'Done', 'Blocked'],
                                 title='توزیع فعالیت‌ها بر اساس سطح سازمانی',
                                 barmode='stack',
                                 labels={'value': 'تعداد فعالیت', 'variable': 'وضعیت'})
                    st.plotly_chart(fig4, use_container_width=True)
                    
                else:
                    st.info(f"هیچ فعالیتی برای {selected_person} یافت نشد.")
        else:
            st.info("لطفاً ابتدا پرسنل را به چارت اضافه کنید.")

# ================================================
# توابع پیشرفته تولید گزارش با فرمت‌بندی Excel
# ================================================

def create_styled_excel_report(df, sheet_name="گزارش فعالیت‌ها", title="گزارش عملکرد"):
    """ایجاد فایل Excel با فرمت‌بندی زیبا و رنگ‌بندی شرطی"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # تعریف استایل‌ها
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        title_font = Font(bold=True, size=14, color='1F4E78')
        
        # استایل برای ردیف‌های شرطی
        todo_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # نارنجی کمرنگ
        doing_fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6', fill_type='solid')  # آبی کمرنگ
        overdue_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # صورتی (عقب‌افتاده)
        done_fill = PatternFill(start_color='C0F0C0', end_color='C0F0C0', fill_type='solid')  # سبز کمرنگ
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تنظیم عرض ستون‌ها
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width
        
        # عنوان گزارش
        worksheet.merge_cells(f'A1:{chr(64+len(df.columns))}1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تاریخ تولید گزارش
        worksheet.merge_cells(f'A2:{chr(64+len(df.columns))}2')
        date_cell = worksheet['A2']
        date_cell.value = f"تاریخ تولید گزارش: {datetime.now().strftime('%Y/%m/%d %H:%M')}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = Alignment(horizontal='right')
        
        # استایل هدر
        for col_num, value in enumerate(df.columns.values, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # رنگ‌بندی بر اساس وضعیت
        if 'وضعیت' in df.columns:
            status_col = df.columns.get_loc('وضعیت') + 1
            for row_num, (_, row_data) in enumerate(df.iterrows(), 4):
                status = row_data.get('وضعیت', '')
                
                # تعیین رنگ بر اساس وضعیت و تاریخ
                fill_color = None
                
                if status == 'ToDo':
                    fill_color = todo_fill
                elif status == 'Doing':
                    fill_color = doing_fill
                elif status == 'Done':
                    fill_color = done_fill
                
                # بررسی عقب‌افتادگی از تاریخ پایان
                if 'تاریخ پایان' in df.columns and status != 'Done':
                    end_date = row_data.get('تاریخ پایان')
                    if pd.notna(end_date) and end_date:
                        try:
                            end_date = pd.to_datetime(end_date)
                            if end_date.date() < datetime.now().date():
                                fill_color = overdue_fill
                        except:
                            pass
                
                # اعمال رنگ به تمام سلول‌های ردیف
                if fill_color:
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.fill = fill_color
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center' if col_num != 2 else 'left', 
                                                  vertical='center', wrap_text=True)
        
        # افزودن فیلتر خودکار
        worksheet.auto_filter.ref = f'A3:{chr(64+len(df.columns))}3'
        
        # ثابت کردن پنجره
        worksheet.freeze_panes = 'A4'
    
    output.seek(0)
    return output

def generate_smart_personnel_report(df, responsible, report_type="daily"):
    """
    تولید گزارش هوشمند برای پرسنل با قابلیت‌های:
    - فقط فعالیت‌های ToDo و Doing
    - تشخیص فعالیت‌های عقب‌افتاده
    - تشخیص فعالیت‌های موعددار در هفته جاری
    """
    # دریافت فعالیت‌های مربوط به این پرسنل
    mask = (df['مسئول1'] == responsible) | (df['مسئول2'] == responsible)
    all_activities = df[mask].copy()
    
    if all_activities.empty:
        return None
    
    # فیلتر بر اساس وضعیت (فقط ToDo و Doing)
    active_activities = all_activities[all_activities['وضعیت'].isin(['ToDo', 'Doing'])]
    
    if active_activities.empty:
        return None
    
    # ستون‌های مورد نیاز
    report_columns = ['ردیف', 'فعالیت', 'وضعیت', 'اولویت_محاسبه‌شده', 
                      'درصد پیشرفت واقعی', 'تاریخ شروع', 'تاریخ پایان', 'توضیحات']
    
    available_columns = [col for col in report_columns if col in active_activities.columns]
    report_df = active_activities[available_columns].copy()
    
    # اضافه کردن ستون‌های تحلیلی
    today = datetime.now().date()
    
    # وضعیت زمانی
    def get_time_status(row):
        if row.get('وضعیت') == 'Done':
            return 'تکمیل شده'
        
        end_date = row.get('تاریخ پایان')
        if pd.isna(end_date) or not end_date:
            return 'بدون تاریخ پایان'
        
        try:
            end_date = pd.to_datetime(end_date).date()
            days_remaining = (end_date - today).days
            
            if days_remaining < 0:
                return f"عقب‌افتاده ({abs(days_remaining)} روز)"
            elif days_remaining == 0:
                return "آخرین مهلت (امروز)"
            elif days_remaining <= 7:
                return f"موعددار این هفته ({days_remaining} روز)"
            else:
                return f"در برنامه ({days_remaining} روز)"
        except:
            return 'تاریخ نامعتبر'
    
    report_df['وضعیت زمانی'] = report_df.apply(get_time_status, axis=1)
    
    # اولویت بندی برای انجام
    priority_map = {'حیاتی': 1, 'بسیار مهم': 2, 'مهم': 3, 'قابل بررسی': 4, 'نامشخص': 5}
    report_df['اولویت عددی'] = report_df['اولویت_محاسبه‌شده'].map(priority_map).fillna(5)
    
    # مرتب‌سازی بر اساس اولویت و وضعیت زمانی
    report_df = report_df.sort_values(['اولویت عددی', 'وضعیت زمانی'], ascending=[True, True])
    report_df = report_df.drop('اولویت عددی', axis=1)
    
    # اضافه کردن ستون‌های عملیاتی
    report_df['ساعت_شروع'] = ''
    report_df['ساعت_پایان'] = ''
    report_df['یادداشت روزانه'] = ''
    
    return report_df

def generate_weekly_overdue_report(df):
    """تولید گزارش هفتگی فعالیت‌های عقب‌افتاده و موعددار"""
    today = datetime.now().date()
    next_week = today + timedelta(days=7)
    
    overdue_activities = []
    due_this_week = []
    
    for _, row in df.iterrows():
        if row.get('وضعیت') in ['Done', 'Blocked']:
            continue
            
        end_date = row.get('تاریخ پایان')
        if pd.isna(end_date) or not end_date:
            continue
            
        try:
            end_date = pd.to_datetime(end_date).date()
            
            if end_date < today:
                # عقب‌افتاده
                overdue_activities.append({
                    'ردیف': row.get('ردیف'),
                    'فعالیت': row.get('فعالیت'),
                    'مسئول1': row.get('مسئول1'),
                    'مسئول2': row.get('مسئول2'),
                    'پوزیشن': row.get('پوزیشن_سازمانی'),
                    'وضعیت': row.get('وضعیت'),
                    'اولویت': row.get('اولویت_محاسبه‌شده'),
                    'تاریخ پایان': end_date,
                    'تعداد روز عقب‌افتادگی': (today - end_date).days,
                    'پیشرفت فعلی': row.get('درصد پیشرفت واقعی', 0)
                })
            elif end_date <= next_week:
                # موعددار این هفته
                due_this_week.append({
                    'ردیف': row.get('ردیف'),
                    'فعالیت': row.get('فعالیت'),
                    'مسئول1': row.get('مسئول1'),
                    'مسئول2': row.get('مسئول2'),
                    'پوزیشن': row.get('پوزیشن_سازمانی'),
                    'وضعیت': row.get('وضعیت'),
                    'اولویت': row.get('اولویت_محاسبه‌شده'),
                    'تاریخ پایان': end_date,
                    'روزهای باقیمانده': (end_date - today).days,
                    'پیشرفت فعلی': row.get('درصد پیشرفت واقعی', 0)
                })
        except:
            continue
    
    return pd.DataFrame(overdue_activities), pd.DataFrame(due_this_week)

# ================================================
# توابع خروجی گزارش‌های BI
# ================================================

def export_bi_report(df, history_df, chart_data, report_type="monthly"):
    """تولید گزارش جامع BI برای ارائه به مدیریت"""
    
    # ایجاد فایل خروجی
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        # 1. برگه خلاصه اجرایی
        executive_summary = pd.DataFrame({
            'شاخص': [
                'تاریخ گزارش',
                'تعداد کل فعالیت‌ها',
                'فعالیت‌های در حال انجام',
                'فعالیت‌های تکمیل شده',
                'فعالیت‌های حیاتی',
                'میانگین پیشرفت کل',
                'درصد تکمیل کل',
                'تعداد پرسنل فعال',
                'تعداد پوزیشن‌های فعال'
            ],
            'مقدار': [
                current_date,
                len(df),
                len(df[df['وضعیت'] == 'Doing']),
                len(df[df['وضعیت'] == 'Done']),
                len(df[df['اولویت_محاسبه‌شده'] == 'حیاتی']),
                f"{df['درصد پیشرفت واقعی'].mean():.1f}%",
                f"{(len(df[df['وضعیت'] == 'Done']) / len(df) * 100):.1f}%",
                len(get_unique_responsibles(df)),
                df['پوزیشن_سازمانی'].nunique()
            ]
        })
        executive_summary.to_excel(writer, sheet_name='خلاصه اجرایی', index=False)
        
        # 2. برگه تحلیل وضعیت
        status_analysis = df['وضعیت'].value_counts().reset_index()
        status_analysis.columns = ['وضعیت', 'تعداد']
        status_analysis['درصد'] = (status_analysis['تعداد'] / len(df) * 100).round(1)
        status_analysis.to_excel(writer, sheet_name='تحلیل وضعیت', index=False)
        
        # 3. برگه تحلیل اولویت
        priority_analysis = df['اولویت_محاسبه‌شده'].value_counts().reset_index()
        priority_analysis.columns = ['اولویت', 'تعداد']
        priority_analysis.to_excel(writer, sheet_name='تحلیل اولویت', index=False)
        
        # 4. برگه تحلیل پرسنل
        personnel_data = []
        for person in chart_data['personnel']:
            person_acts = get_activities_for_responsible(df, person)
            if not person_acts.empty:
                personnel_data.append({
                    'نام پرسنل': person,
                    'تعداد کل فعالیت': len(person_acts),
                    'ToDo': len(person_acts[person_acts['وضعیت'] == 'ToDo']),
                    'Doing': len(person_acts[person_acts['وضعیت'] == 'Doing']),
                    'Done': len(person_acts[person_acts['وضعیت'] == 'Done']),
                    'Blocked': len(person_acts[person_acts['وضعیت'] == 'Blocked']),
                    'میانگین پیشرفت': person_acts['درصد پیشرفت واقعی'].mean(),
                    'تعداد پوزیشن': len(chart_data['personnel_positions'].get(person, []))
                })
        if personnel_data:
            personnel_df = pd.DataFrame(personnel_data)
            personnel_df = personnel_df.sort_values('تعداد کل فعالیت', ascending=False)
            personnel_df.to_excel(writer, sheet_name='تحلیل پرسنل', index=False)
        
        # 5. برگه تحلیل پوزیشن
        position_data = []
        for position in chart_data['positions']:
            pos_acts = df[df['پوزیشن_سازمانی'] == position]
            if not pos_acts.empty:
                position_data.append({
                    'پوزیشن': position,
                    'تعداد کل فعالیت': len(pos_acts),
                    'ToDo': len(pos_acts[pos_acts['وضعیت'] == 'ToDo']),
                    'Doing': len(pos_acts[pos_acts['وضعیت'] == 'Doing']),
                    'Done': len(pos_acts[pos_acts['وضعیت'] == 'Done']),
                    'Blocked': len(pos_acts[pos_acts['وضعیت'] == 'Blocked']),
                    'میانگین پیشرفت': pos_acts['درصد پیشرفت واقعی'].mean(),
                    'تعداد پرسنل فعال': len(set(pos_acts['مسئول1'].dropna()) | set(pos_acts['مسئول2'].dropna()))
                })
        if position_data:
            position_df = pd.DataFrame(position_data)
            position_df = position_df.sort_values('تعداد کل فعالیت', ascending=False)
            position_df.to_excel(writer, sheet_name='تحلیل پوزیشن', index=False)
        
        # 6. برگه فعالیت‌های حیاتی
        critical_acts = df[df['اولویت_محاسبه‌شده'] == 'حیاتی'].copy()
        critical_cols = ['ردیف', 'فعالیت', 'وضعیت', 'مسئول1', 'مسئول2', 
                        'پوزیشن_سازمانی', 'درصد پیشرفت واقعی', 'تاریخ پایان']
        available_critical = [col for col in critical_cols if col in critical_acts.columns]
        if not critical_acts.empty:
            critical_acts[available_critical].to_excel(writer, sheet_name='فعالیت‌های حیاتی', index=False)
        
        # 7. برگه فعالیت‌های عقب‌افتاده
        overdue_df, due_df = generate_weekly_overdue_report(df)
        if not overdue_df.empty:
            overdue_df.to_excel(writer, sheet_name='عقب‌افتاده', index=False)
        if not due_df.empty:
            due_df.to_excel(writer, sheet_name='موعددار این هفته', index=False)
        
        # 8. برگه روند زمانی (اگر تاریخچه وجود دارد)
        if not history_df.empty:
            history_clean = prepare_history_for_analysis(history_df)
            if not history_clean.empty and 'تاریخ' in history_clean.columns:
                history_clean['تاریخ'] = pd.to_datetime(history_clean['تاریخ'])
                history_clean['ماه'] = history_clean['تاریخ'].dt.to_period('M').astype(str)
                
                trend_data = history_clean.groupby('ماه').size().reset_index(name='تعداد تغییرات')
                trend_data.to_excel(writer, sheet_name='روند زمانی', index=False)
        
        # اعمال فرمت‌بندی
        workbook = writer.book
        
        # تنظیم عرض ستون‌ها برای همه برگه‌ها
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width
    
    output.seek(0)
    return output

def prepare_history_for_analysis(history_df):
    """آماده‌سازی تاریخچه برای تحلیل"""
    if history_df.empty:
        return pd.DataFrame()
    
    # کپی از دیتافریم
    df = history_df.copy()
    
    # تبدیل ستون تاریخ به datetime
    if 'تاریخ' in df.columns:
        df['تاریخ'] = pd.to_datetime(df['تاریخ'], errors='coerce')
        # حذف ردیف‌هایی که تاریخ نامعتبر دارند
        df = df.dropna(subset=['تاریخ'])
    
    return df

def analyze_positions_status(df, chart_data):
    """تحلیل وضعیت پوزیشن‌ها بر اساس فعالیت‌ها"""
    
    # استخراج تمام پوزیشن‌ها
    all_positions = extract_all_positions(chart_data.get('structure', create_organizational_structure()))
    
    position_status = []
    
    for _, pos_row in all_positions.iterrows():
        position_name = pos_row['نام']
        position_activities = df[df['پوزیشن_سازمانی'] == position_name]
        
        if not position_activities.empty:
            status = {
                'پوزیشن': position_name,
                'مسیر کامل': pos_row['مسیر کامل'],
                'سطح': pos_row['سطح'],
                'نوع': pos_row['نوع'],
                'تعداد کل': len(position_activities),
                'ToDo': (position_activities['وضعیت'] == 'ToDo').sum(),
                'Doing': (position_activities['وضعیت'] == 'Doing').sum(),
                'Done': (position_activities['وضعیت'] == 'Done').sum(),
                'Blocked': (position_activities['وضعیت'] == 'Blocked').sum(),
                'درصد تکمیل': round((position_activities['وضعیت'] == 'Done').sum() / len(position_activities) * 100, 1),
                'میانگین پیشرفت': round(position_activities['درصد پیشرفت واقعی'].mean(), 1)
            }
            position_status.append(status)
    
    return pd.DataFrame(position_status)

def advanced_bi_analysis(df, history_df, chart_data):
    """تحلیل‌های پیشرفته BI - با ساختار سلسله‌مراتبی پوزیشن‌ها"""
    
    st.header("📊 داشبورد مدیریتی BI")
    
    if df.empty:
        st.warning("داده‌ای برای تحلیل وجود ندارد!")
        return
    
    # آماده‌سازی تاریخچه
    history_clean = prepare_history_for_analysis(history_df)
    
    # تحلیل وضعیت پوزیشن‌ها
    position_status_df = analyze_positions_status(df, chart_data)
    
    # ایجاد تب‌های تحلیلی
    bi_tab1, bi_tab2, bi_tab3, bi_tab4, bi_tab5, bi_tab6, bi_tab7 = st.tabs([
        "📈 توزیع فعالیت‌ها", 
        "👥 عملکرد پرسنل",
        "🏢 تحلیل پوزیشن‌ها",
        "📊 ماتریس پوزیشن-وضعیت",
        "🌳 سلسله‌مراتب سازمانی",
        "⏱️ تحلیل زمانی",
        "📉 روندها و پیش‌بینی"
    ])
    
    # ========================================
    # تب 1: توزیع فعالیت‌ها
    # ========================================
    with bi_tab1:
        st.subheader("توزیع فعالیت‌ها بر اساس معیارهای مختلف")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # توزیع بر اساس وضعیت
            if 'وضعیت' in df.columns:
                status_dist = df['وضعیت'].value_counts().reset_index()
                status_dist.columns = ['وضعیت', 'تعداد']
                
                fig1 = px.pie(status_dist, values='تعداد', names='وضعیت', 
                             title='توزیع فعالیت‌ها بر اساس وضعیت',
                             color_discrete_map={
                                 'ToDo': '#FFA07A',
                                 'Doing': '#87CEEB',
                                 'Done': '#98FB98',
                                 'Blocked': '#FF6347'
                             })
                st.plotly_chart(fig1, use_container_width=True)
        
        with col2:
            # توزیع بر اساس اولویت
            if 'اولویت_محاسبه‌شده' in df.columns:
                priority_dist = df['اولویت_محاسبه‌شده'].value_counts().reset_index()
                priority_dist.columns = ['اولویت', 'تعداد']
                
                fig2 = px.bar(priority_dist, x='اولویت', y='تعداد',
                             title='توزیع فعالیت‌ها بر اساس اولویت',
                             color='اولویت', text='تعداد')
                st.plotly_chart(fig2, use_container_width=True)
    
    # ========================================
    # تب 2: عملکرد پرسنل
    # ========================================
    with bi_tab2:
        st.subheader("تحلیل عملکرد پرسنل")
        
        # آماده‌سازی داده‌های پرسنلی
        all_personnel = []
        for _, row in df.iterrows():
            if row.get('مسئول1') and row['مسئول1'] != '':
                all_personnel.append({
                    'نام': row['مسئول1'],
                    'پوزیشن': row.get('پوزیشن_سازمانی', 'نامشخص'),
                    'وضعیت': row.get('وضعیت', ''),
                    'اولویت': row.get('اولویت_محاسبه‌شده', ''),
                    'پیشرفت': row.get('درصد پیشرفت واقعی', 0)
                })
            if row.get('مسئول2') and row['مسئول2'] != '':
                all_personnel.append({
                    'نام': row['مسئول2'],
                    'پوزیشن': row.get('پوزیشن_سازمانی', 'نامشخص'),
                    'وضعیت': row.get('وضعیت', ''),
                    'اولویت': row.get('اولویت_محاسبه‌شده', ''),
                    'پیشرفت': row.get('درصد پیشرفت واقعی', 0)
                })
        
        personnel_df = pd.DataFrame(all_personnel)
        
        if not personnel_df.empty:
            # آمار تجمعی پرسنل
            personnel_stats = []
            for name in personnel_df['نام'].unique():
                person_data = personnel_df[personnel_df['نام'] == name]
                stats = {
                    'نام': name,
                    'تعداد کل فعالیت‌ها': len(person_data),
                    'میانگین پیشرفت': person_data['پیشرفت'].mean(),
                    'تعداد Done': (person_data['وضعیت'] == 'Done').sum(),
                    'تعداد فعالیت‌های حیاتی': (person_data['اولویت'] == 'حیاتی').sum(),
                    'تعداد پوزیشن‌های متفاوت': person_data['پوزیشن'].nunique()
                }
                personnel_stats.append(stats)
            
            personnel_stats_df = pd.DataFrame(personnel_stats)
            personnel_stats_df = personnel_stats_df.round(2)
            
            st.dataframe(personnel_stats_df, use_container_width=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                fig3 = px.bar(personnel_stats_df.sort_values('تعداد کل فعالیت‌ها', ascending=False).head(10), 
                             x='نام', y='تعداد کل فعالیت‌ها',
                             title='۱۰ پرسنل با بیشترین فعالیت',
                             color='نام', text='تعداد کل فعالیت‌ها')
                st.plotly_chart(fig3, use_container_width=True)
            
            with col2:
                fig4 = px.bar(personnel_stats_df.sort_values('میانگین پیشرفت', ascending=False).head(10),
                             x='نام', y='میانگین پیشرفت',
                             title='۱۰ پرسنل با بالاترین میانگین پیشرفت',
                             color='نام', text='میانگین پیشرفت')
                st.plotly_chart(fig4, use_container_width=True)
    
    # ========================================
    # تب 3: تحلیل پوزیشن‌ها
    # ========================================
    with bi_tab3:
        st.subheader("تحلیل پوزیشن‌های سازمانی")
        
        if not position_status_df.empty:
            # فیلتر بر اساس سطح
            levels = sorted(position_status_df['سطح'].unique())
            selected_levels = st.multiselect(
                "فیلتر بر اساس سطح سازمانی",
                options=levels,
                default=levels
            )
            
            filtered_positions = position_status_df[position_status_df['سطح'].isin(selected_levels)]
            
            st.dataframe(filtered_positions, use_container_width=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # ۱۰ پوزیشن با بیشترین فعالیت
                top_positions = filtered_positions.sort_values('تعداد کل', ascending=False).head(10)
                fig5 = px.bar(top_positions, x='پوزیشن', y='تعداد کل',
                             title='۱۰ پوزیشن با بیشترین فعالیت',
                             color='پوزیشن', text='تعداد کل')
                st.plotly_chart(fig5, use_container_width=True)
            
            with col2:
                # توزیع وضعیت در پوزیشن‌ها
                status_by_position = filtered_positions.melt(
                    id_vars=['پوزیشن'], 
                    value_vars=['ToDo', 'Doing', 'Done', 'Blocked'],
                    var_name='وضعیت', 
                    value_name='تعداد'
                )
                fig6 = px.bar(status_by_position, x='پوزیشن', y='تعداد', color='وضعیت',
                             title='توزیع وضعیت فعالیت‌ها در پوزیشن‌ها',
                             barmode='stack',
                             color_discrete_map={
                                 'ToDo': '#FFA07A',
                                 'Doing': '#87CEEB',
                                 'Done': '#98FB98',
                                 'Blocked': '#FF6347'
                             })
                st.plotly_chart(fig6, use_container_width=True)
    
    # ========================================
    # تب 4: ماتریس پوزیشن-وضعیت
    # ========================================
    with bi_tab4:
        st.subheader("ماتریس توزیع وضعیت در پوزیشن‌ها")
        
        if not position_status_df.empty:
            # ایجاد ماتریس حرارتی
            pivot_matrix = position_status_df.pivot_table(
                values=['ToDo', 'Doing', 'Done', 'Blocked'],
                index='پوزیشن'
            ).head(20)  # محدود به 20 پوزیشن اول
            
            fig7 = px.imshow(pivot_matrix.T,
                            title='ماتریس حرارتی توزیع وضعیت در پوزیشن‌ها',
                            labels=dict(x="پوزیشن", y="وضعیت", color="تعداد"),
                            aspect="auto",
                            color_continuous_scale='Viridis')
            st.plotly_chart(fig7, use_container_width=True)
            
            # نمودار درصد تکمیل به تفکیک سطح
            completion_by_level = position_status_df.groupby('سطح').agg({
                'تعداد کل': 'sum',
                'Done': 'sum'
            }).reset_index()
            completion_by_level['درصد تکمیل'] = (completion_by_level['Done'] / completion_by_level['تعداد کل'] * 100).round(1)
            
            fig8 = px.bar(completion_by_level, x='سطح', y='درصد تکمیل',
                         title='درصد فعالیت‌های تکمیل شده به تفکیک سطح سازمانی',
                         color='سطح', text='درصد تکمیل')
            st.plotly_chart(fig8, use_container_width=True)
    
    # ========================================
    # تب 5: سلسله‌مراتب سازمانی
    # ========================================
    with bi_tab5:
        st.subheader("نمایش سلسله‌مراتبی وضعیت پوزیشن‌ها")
        
        # ایجاد نمودار درختی
        import plotly.figure_factory as ff
        
        # آماده‌سازی داده‌ها برای نمودار درختی
        structure = chart_data.get('structure', create_organizational_structure())
        
        def prepare_treemap_data(node, parent="", path=""):
            data = []
            if isinstance(node, dict):
                for key, value in node.items():
                    current_path = f"{path}/{key}" if path else key
                    
                    # پیدا کردن آمار این پوزیشن
                    pos_stats = position_status_df[position_status_df['پوزیشن'] == key]
                    if not pos_stats.empty:
                        total = pos_stats.iloc[0]['تعداد کل']
                        done = pos_stats.iloc[0]['Done']
                    else:
                        total = 0
                        done = 0
                    
                    data.append({
                        'id': current_path,
                        'parent': parent,
                        'name': key,
                        'value': total if total > 0 else 1,
                        'completed': done,
                        'total': total
                    })
                    
                    data.extend(prepare_treemap_data(value, current_path, current_path))
            elif isinstance(node, list):
                for item in node:
                    current_path = f"{path}/{item}" if path else item
                    
                    pos_stats = position_status_df[position_status_df['پوزیشن'] == item]
                    if not pos_stats.empty:
                        total = pos_stats.iloc[0]['تعداد کل']
                        done = pos_stats.iloc[0]['Done']
                    else:
                        total = 0
                        done = 0
                    
                    data.append({
                        'id': current_path,
                        'parent': parent,
                        'name': item,
                        'value': total if total > 0 else 1,
                        'completed': done,
                        'total': total
                    })
            return data
        
        treemap_data = prepare_treemap_data(structure)
        
        if treemap_data:
            treemap_df = pd.DataFrame(treemap_data)
            treemap_df = treemap_df[treemap_df['parent'] != ""]
            
            fig9 = px.treemap(
                treemap_df,
                ids='id',
                parents='parent',
                names='name',
                values='value',
                color='completed',
                color_continuous_scale='RdYlGn',
                title='سلسله‌مراتب سازمانی با رنگ‌بندی بر اساس تعداد فعالیت‌های تکمیل شده',
                hover_data=['total']
            )
            st.plotly_chart(fig9, use_container_width=True)
    
    # ========================================
    # تب 6: تحلیل زمانی
    # ========================================
    with bi_tab6:
        st.subheader("تحلیل‌های زمانی و Lead Time")
        
        if not history_clean.empty:
            # تحلیل مدت زمان انجام فعالیت‌ها
            lead_times = []
            for _, row in df.iterrows():
                if row.get('وضعیت') == 'Done':
                    activity_history = get_activity_history(row.get('ردیف', ''))
                    lead_time = calculate_lead_time(activity_history, row)
                    if lead_time is not None:
                        lead_times.append({
                            'پوزیشن': row.get('پوزیشن_سازمانی', 'نامشخص'),
                            'مدت_زمان_(روز)': lead_time,
                            'اولویت': row.get('اولویت_محاسبه‌شده', '')
                        })
            
            if lead_times:
                lead_df = pd.DataFrame(lead_times)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # میانگین زمان انجام به تفکیک پوزیشن
                    avg_lead_pos = lead_df.groupby('پوزیشن')['مدت_زمان_(روز)'].mean().reset_index()
                    avg_lead_pos = avg_lead_pos.sort_values('مدت_زمان_(روز)', ascending=False).head(10)
                    fig10 = px.bar(avg_lead_pos, x='پوزیشن', y='مدت_زمان_(روز)',
                                 title='میانگین زمان انجام فعالیت به تفکیک پوزیشن (۱۰ پوزیشن برتر)',
                                 color='پوزیشن', text='مدت_زمان_(روز)')
                    st.plotly_chart(fig10, use_container_width=True)
    
    # ========================================
    # تب 7: روندها و پیش‌بینی
    # ========================================
    with bi_tab7:
        st.subheader("تحلیل روندها و پیش‌بینی")
        
        if not history_clean.empty:
            # روند تکمیل فعالیت‌ها
            if 'وضعیت_جدید' in history_clean.columns:
                completion_data = history_clean[history_clean['وضعیت_جدید'] == 'Done'].copy()
                if not completion_data.empty:
                    completion_data['ماه'] = completion_data['تاریخ'].dt.to_period('M').astype(str)
                    completion_trend = completion_data.groupby('ماه').size().reset_index(name='تعداد_تکمیل_شده')
                    
                    fig11 = px.line(completion_trend, x='ماه', y='تعداد_تکمیل_شده',
                                   title='روند ماهانه تکمیل فعالیت‌ها',
                                   markers=True)
                    st.plotly_chart(fig11, use_container_width=True)
        
        # تحلیل بار کاری جاری
        st.subheader("تحلیل بار کاری جاری")
        
        if 'وضعیت' in df.columns:
            doing_activities = df[df['وضعیت'] == 'Doing']
            if not doing_activities.empty:
                workload_by_position = doing_activities['پوزیشن_سازمانی'].value_counts().reset_index()
                workload_by_position.columns = ['پوزیشن', 'تعداد_در_حال_انجام']
                workload_by_position = workload_by_position.head(10)
                
                fig12 = px.bar(workload_by_position, x='پوزیشن', y='تعداد_در_حال_انجام',
                             title='بار کاری جاری به تفکیک پوزیشن (۱۰ پوزیشن برتر)',
                             color='پوزیشن', text='تعداد_در_حال_انجام')
                st.plotly_chart(fig12, use_container_width=True)

# ================================================
# توابع مدیریت فیلتر سراسری
# ================================================

def initialize_session_state():
    """مقداردهی اولیه session state برای فیلترهای سراسری"""
    if 'global_responsibles' not in st.session_state:
        st.session_state.global_responsibles = []
    if 'global_statuses' not in st.session_state:
        st.session_state.global_statuses = []
    if 'global_priorities' not in st.session_state:
        st.session_state.global_priorities = []
    if 'global_progress_range' not in st.session_state:
        st.session_state.global_progress_range = (0, 100)

def apply_global_filters(df):
    """اعمال فیلترهای سراسری روی دیتافریم"""
    filtered_df = df.copy()
    
    if st.session_state.global_responsibles:
        mask = False
        for resp in st.session_state.global_responsibles:
            mask |= (filtered_df['مسئول1'] == resp) | (filtered_df['مسئول2'] == resp)
        filtered_df = filtered_df[mask]
    
    if st.session_state.global_statuses:
        filtered_df = filtered_df[filtered_df['وضعیت'].isin(st.session_state.global_statuses)]
    
    if st.session_state.global_priorities:
        filtered_df = filtered_df[filtered_df['اولویت_محاسبه‌شده'].isin(st.session_state.global_priorities)]
    
    progress_min, progress_max = st.session_state.global_progress_range
    filtered_df = filtered_df[
        (filtered_df['درصد پیشرفت واقعی'] >= progress_min) & 
        (filtered_df['درصد پیشرفت واقعی'] <= progress_max)
    ]
    
    return filtered_df

def render_global_filters(df, chart_data):
    """نمایش فیلترهای سراسری در سایدبار"""
    with st.sidebar:
        st.header("🔍 فیلترهای سراسری")
        
        # فیلتر چندگانه مسئولین
        selected_responsibles = st.multiselect(
            "مسئولین",
            options=chart_data['personnel'],
            default=st.session_state.global_responsibles,
            key="global_resp_filter"
        )
        st.session_state.global_responsibles = selected_responsibles
        
        # فیلتر چندگانه وضعیت
        all_statuses = [s for s in df['وضعیت'].dropna().unique() if s]
        selected_statuses = st.multiselect(
            "وضعیت",
            options=all_statuses,
            default=st.session_state.global_statuses,
            key="global_status_filter"
        )
        st.session_state.global_statuses = selected_statuses
        
        # فیلتر چندگانه اولویت
        all_priorities = ['حیاتی', 'بسیار مهم', 'مهم', 'قابل بررسی', 'نامشخص']
        selected_priorities = st.multiselect(
            "اولویت",
            options=all_priorities,
            default=st.session_state.global_priorities,
            key="global_priority_filter"
        )
        st.session_state.global_priorities = selected_priorities
        
        # فیلتر بازه پیشرفت
        progress_range = st.slider(
            "محدوده درصد پیشرفت",
            min_value=0, max_value=100,
            value=st.session_state.global_progress_range,
            key="global_progress_filter"
        )
        st.session_state.global_progress_range = progress_range
        
        st.markdown("---")
        if st.button("🔄 بازنشانی همه فیلترها"):
            st.session_state.global_responsibles = []
            st.session_state.global_statuses = []
            st.session_state.global_priorities = []
            st.session_state.global_progress_range = (0, 100)
            st.rerun()

def render_reports_tab(df, chart_data):
    """تب تولید گزارش با قابلیت‌های جدید"""
    st.header("📊 تولید گزارش‌های پیشرفته")
    
    report_type = st.radio(
        "نوع گزارش",
        ["📋 گزارش روزانه پرسنل", "📈 گزارش هفتگی عملکرد", "📑 گزارش جامع مدیریتی BI"],
        horizontal=True
    )
    
    if report_type == "📋 گزارش روزانه پرسنل":
        st.subheader("گزارش وظایف روزانه پرسنل")
        st.info("این گزارش فقط فعالیت‌های با وضعیت ToDo و Doing را شامل می‌شود.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # انتخاب نوع گزارش
            report_scope = st.radio(
                "محدوده گزارش",
                ["همه پرسنل", "پرسنل مشخص", "فعالیت‌های بدون مسئول"]
            )
            
            if report_scope == "پرسنل مشخص":
                selected_personnel = st.selectbox(
                    "انتخاب پرسنل", 
                    chart_data['personnel'],
                    key="report_person_select"
                )
        
        with col2:
            st.markdown("### تنظیمات پیشرفته")
            include_overdue = st.checkbox("🟥 مشخص کردن فعالیت‌های عقب‌افتاده", value=True)
            include_priority = st.checkbox("⭐ مرتب‌سازی بر اساس اولویت", value=True)
            color_format = st.checkbox("🎨 اعمال رنگ‌بندی خودکار", value=True)
        
        if st.button("📥 تولید گزارش", type="primary", key="generate_daily_report"):
            with st.spinner("در حال تولید گزارش..."):
                try:
                    if report_scope == "همه پرسنل":
                        # تولید گزارش برای همه پرسنل
                        reports_generated = []
                        for person in chart_data['personnel']:
                            report_df = generate_smart_personnel_report(df, person)
                            if report_df is not None:
                                filename = f"reports/Daily_{person}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                                report_path = Path('reports') / filename
                                
                                # ایجاد فایل با فرمت‌بندی
                                excel_output = create_styled_excel_report(
                                    report_df, 
                                    sheet_name=f"وظایف {person}",
                                    title=f"گزارش وظایف روزانه - {person}"
                                )
                                
                                with open(report_path, 'wb') as f:
                                    f.write(excel_output.getvalue())
                                
                                reports_generated.append(filename)
                        
                        if reports_generated:
                            st.success(f"✅ {len(reports_generated)} فایل گزارش با موفقیت تولید شد!")
                            
                            # نمایش لیست فایل‌ها
                            with st.expander("📂 مشاهده لیست فایل‌های تولید شده"):
                                for f in reports_generated:
                                    st.write(f"- {f}")
                        else:
                            st.warning("هیچ فعالیت فعالی برای پرسنل یافت نشد!")
                    
                    elif report_scope == "پرسنل مشخص":
                        report_df = generate_smart_personnel_report(df, selected_personnel)
                        
                        if report_df is not None:
                            filename = f"reports/Daily_{selected_personnel}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                            report_path = Path('reports') / filename
                            
                            excel_output = create_styled_excel_report(
                                report_df,
                                sheet_name=f"وظایف {selected_personnel}",
                                title=f"گزارش وظایف روزانه - {selected_personnel}"
                            )
                            
                            with open(report_path, 'wb') as f:
                                f.write(excel_output.getvalue())
                            
                            st.success(f"✅ گزارش برای {selected_personnel} تولید شد!")
                            
                            # نمایش پیش‌نمایش
                            with st.expander("👁️ پیش‌نمایش گزارش"):
                                st.dataframe(report_df, use_container_width=True)
                            
                            # دکمه دانلود
                            with open(report_path, 'rb') as f:
                                st.download_button(
                                    label="📥 دانلود فایل Excel",
                                    data=f,
                                    file_name=filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        else:
                            st.warning(f"هیچ فعالیت فعالی برای {selected_personnel} یافت نشد!")
                    
                    else:  # فعالیت‌های بدون مسئول
                        without_owner = get_activities_without_responsible(df)
                        without_owner_active = without_owner[without_owner['وضعیت'].isin(['ToDo', 'Doing'])]
                        
                        if not without_owner_active.empty:
                            filename = f"reports/WithoutOwner_{datetime.now().strftime('%Y%m%d')}.xlsx"
                            report_path = Path('without_owner') / filename
                            
                            excel_output = create_styled_excel_report(
                                without_owner_active,
                                sheet_name="بدون مسئول",
                                title="گزارش فعالیت‌های بدون مسئول"
                            )
                            
                            with open(report_path, 'wb') as f:
                                f.write(excel_output.getvalue())
                            
                            st.success(f"✅ گزارش فعالیت‌های بدون مسئول تولید شد!")
                        else:
                            st.warning("هیچ فعالیت بدون مسئول فعالی یافت نشد!")
                            
                except Exception as e:
                    st.error(f"خطا در تولید گزارش: {e}")
    
    elif report_type == "📈 گزارش هفتگی عملکرد":
        st.subheader("گزارش هفتگی عملکرد - فعالیت‌های عقب‌افتاده و موعددار")
        
        # تحلیل هفتگی
        overdue_df, due_df = generate_weekly_overdue_report(df)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("فعالیت‌های عقب‌افتاده", len(overdue_df))
        
        with col2:
            st.metric("فعالیت‌های موعددار این هفته", len(due_df))
        
        if not overdue_df.empty:
            with st.expander("⚠️ فعالیت‌های عقب‌افتاده", expanded=True):
                st.dataframe(overdue_df, use_container_width=True)
        
        if not due_df.empty:
            with st.expander("⏰ فعالیت‌های موعددار این هفته", expanded=True):
                st.dataframe(due_df, use_container_width=True)
        
        if st.button("📥 دریافت گزارش هفتگی", key="weekly_report_btn"):
            # ایجاد گزارش ترکیبی
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                if not overdue_df.empty:
                    # ذخیره موقت و بازخوانی برای فرمت‌بندی
                    temp_file = "temp_overdue.xlsx"
                    overdue_df.to_excel(temp_file, index=False)
                    temp_df = pd.read_excel(temp_file)
                    temp_df.to_excel(writer, sheet_name="عقب‌افتاده", index=False)
                    os.remove(temp_file)
                
                if not due_df.empty:
                    temp_file = "temp_due.xlsx"
                    due_df.to_excel(temp_file, index=False)
                    temp_df = pd.read_excel(temp_file)
                    temp_df.to_excel(writer, sheet_name="موعددار این هفته", index=False)
                    os.remove(temp_file)
            
            output.seek(0)
            
            st.download_button(
                label="📥 دانلود گزارش هفتگی",
                data=output,
                file_name=f"Weekly_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    else:  # گزارش جامع مدیریتی BI
        st.subheader("📑 گزارش جامع مدیریتی BI")
        st.info("این گزارش شامل تمام تحلیل‌های عملکردی برای ارائه به مدیریت است.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            report_period = st.selectbox(
                "دوره گزارش",
                ["هفتگی", "ماهانه", "فصلی", "سالیانه"]
            )
            
            include_charts = st.checkbox("📊 شامل نمودارهای تحلیلی", value=True)
        
        with col2:
            start_date = st.date_input("تاریخ شروع", datetime.now() - timedelta(days=30))
            end_date = st.date_input("تاریخ پایان", datetime.now())
        
        if st.button("📥 تولید گزارش جامع مدیریتی", type="primary", key="bi_report_btn"):
            with st.spinner("در حال تولید گزارش جامع مدیریتی..."):
                try:
                    # بارگذاری تاریخچه
                    history_file = 'history/History.xlsx'
                    history_df = pd.read_excel(history_file) if os.path.exists(history_file) else pd.DataFrame()
                    
                    # تولید گزارش
                    report_output = export_bi_report(df, history_df, chart_data, report_period)
                    
                    # ذخیره فایل
                    filename = f"reports/BI_Report_{report_period}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    report_path = Path('reports') / filename
                    
                    with open(report_path, 'wb') as f:
                        f.write(report_output.getvalue())
                    
                    st.success(f"✅ گزارش جامع مدیریتی با موفقیت تولید شد!")
                    
                    # دکمه دانلود
                    with open(report_path, 'rb') as f:
                        st.download_button(
                            label="📥 دانلود گزارش جامع",
                            data=f,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    # نمایش خلاصه
                    st.balloons()
                    
                except Exception as e:
                    st.error(f"خطا در تولید گزارش: {e}")

# ================================================
# رابط کاربری اصلی
# ================================================

def main():
    st.set_page_config(page_title="سیستم جامع مدیریت فعالیت‌ها", layout="wide")
    
    st.title("🎯 سیستم جامع مدیریت و ارزیابی عملکرد شرکت خوارزمی")
    st.markdown("---")
    
    # ایجاد پوشه‌های مورد نیاز
    setup_directories()
    
    # بارگذاری داده‌ها
    df = load_data()
    if df is None:
        return
    
    # بارگذاری چارت سازمانی
    chart_data = load_organizational_chart()
    
    # به‌روزرسانی چارت از دیتابیس
    update_from_db(df, chart_data)
    
    # مقداردهی اولیه session state
    initialize_session_state()
    
    # نمایش فیلترهای سراسری
    render_global_filters(df, chart_data)
    
    # اعمال فیلترهای سراسری
    filtered_df = apply_global_filters(df)
    
    # بارگذاری تاریخچه
    history_file = 'history/History.xlsx'
    history_df = pd.read_excel(history_file) if os.path.exists(history_file) else pd.DataFrame()
    
    # ایجاد تب‌های اصلی
    main_tabs = st.tabs([
        "📋 داشبورد اصلی",
        "➕ ایجاد فعالیت",
        "✏️ ویرایش فعالیت",
        "🏢 چارت سازمانی",
        "📊 تولید گزارش",
        "📈 تحلیل پیشرفته BI",
        "📚 تاریخچه کامل"
    ])
    
    # ========================================
    # تب 1: داشبورد اصلی
    # ========================================
    with main_tabs[0]:
        st.header("داشبورد اصلی فعالیت‌ها")
        
        # آمار سریع
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("تعداد کل فعالیت‌ها", len(filtered_df))
        with col2:
            st.metric("فعالیت‌های حیاتی", len(filtered_df[filtered_df['اولویت_محاسبه‌شده'] == 'حیاتی']))
        with col3:
            st.metric("در حال انجام", len(filtered_df[filtered_df['وضعیت'] == 'Doing']))
        with col4:
            avg_progress = filtered_df['درصد پیشرفت واقعی'].mean()
            st.metric("میانگین پیشرفت", f"{avg_progress:.1f}%")
        
        # نمایش فعالیت‌ها
        st.subheader("لیست فعالیت‌ها")
        display_columns = ['ردیف', 'فعالیت', 'وضعیت', 'مسئول1', 'مسئول2', 'پوزیشن_سازمانی',
                          'درصد پیشرفت واقعی', 'اولویت_محاسبه‌شده', 'تاریخ شروع', 'تاریخ پایان', 'توضیحات', 'تاریخ_آخرین_تغییر']
        available_display = [col for col in display_columns if col in filtered_df.columns]
        
        display_df = filtered_df[available_display].copy()
        st.dataframe(display_df, use_container_width=True)
    
    # ========================================
    # تب 2: ایجاد فعالیت
    # ========================================
    with main_tabs[1]:
        df = add_new_activity(df, chart_data)
        filtered_df = apply_global_filters(df)
    
    # ========================================
    # تب 3: ویرایش فعالیت
    # ========================================
    with main_tabs[2]:
        df = edit_activity(df, chart_data)
        filtered_df = apply_global_filters(df)
    
    # ========================================
    # تب 4: چارت سازمانی
    # ========================================
    with main_tabs[3]:
        organizational_chart_page(df, chart_data)
    
    # ========================================
    # تب 5: تولید گزارش
    # ========================================
    with main_tabs[4]:
        render_reports_tab(df, chart_data)
    
    # ========================================
    # تب 6: تحلیل پیشرفته BI
    # ========================================
    with main_tabs[5]:
        advanced_bi_analysis(filtered_df, history_df, chart_data)
    
    # ========================================
    # تب 7: تاریخچه کامل
    # ========================================
    with main_tabs[6]:
        st.header("تاریخچه کامل تغییرات")
        
        if not history_df.empty:
            # فیلترهای تاریخچه
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if 'نوع_عملیات' in history_df.columns:
                    op_types = ['همه'] + history_df['نوع_عملیات'].unique().tolist()
                    selected_op = st.selectbox("نوع عملیات", op_types)
                else:
                    selected_op = 'همه'
            
            with col2:
                if 'اولویت' in history_df.columns:
                    priorities_hist = ['همه'] + history_df['اولویت'].unique().tolist()
                    selected_priority_hist = st.selectbox("اولویت", priorities_hist)
                else:
                    selected_priority_hist = 'همه'
            
            with col3:
                if 'پوزیشن' in history_df.columns:
                    positions_hist = ['همه'] + history_df['پوزیشن'].unique().tolist()[:50]
                    selected_position_hist = st.selectbox("پوزیشن", positions_hist)
                else:
                    selected_position_hist = 'همه'
            
            # اعمال فیلتر
            filtered_history = history_df.copy()
            if selected_op != 'همه' and 'نوع_عملیات' in filtered_history.columns:
                filtered_history = filtered_history[filtered_history['نوع_عملیات'] == selected_op]
            if selected_priority_hist != 'همه' and 'اولویت' in filtered_history.columns:
                filtered_history = filtered_history[filtered_history['اولویت'] == selected_priority_hist]
            if selected_position_hist != 'همه' and 'پوزیشن' in filtered_history.columns:
                filtered_history = filtered_history[filtered_history['پوزیشن'] == selected_position_hist]
            
            st.dataframe(filtered_history, use_container_width=True)
            st.markdown(f"**تعداد کل تغییرات:** {len(filtered_history)}")
            
            # آمار تاریخچه
            st.subheader("آمار تاریخچه")
            col1, col2 = st.columns(2)
            
            with col1:
                if 'نوع_عملیات' in filtered_history.columns:
                    op_stats = filtered_history['نوع_عملیات'].value_counts().reset_index()
                    op_stats.columns = ['نوع عملیات', 'تعداد']
                    fig_hist1 = px.pie(op_stats, values='تعداد', names='نوع عملیات',
                                      title='توزیع نوع عملیات‌ها')
                    st.plotly_chart(fig_hist1, use_container_width=True)
            
            with col2:
                if 'پوزیشن' in filtered_history.columns:
                    pos_stats = filtered_history['پوزیشن'].value_counts().reset_index().head(10)
                    pos_stats.columns = ['پوزیشن', 'تعداد']
                    fig_hist2 = px.bar(pos_stats, x='پوزیشن', y='تعداد',
                                      title='۱۰ پوزیشن با بیشترین تغییرات')
                    st.plotly_chart(fig_hist2, use_container_width=True)
        else:
            st.info("تاریخچه‌ای ثبت نشده است.")

# ================================================
# اجرای برنامه
# ================================================
if __name__ == "__main__":
    main()
